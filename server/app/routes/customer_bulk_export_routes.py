"""
Bulk export for the Customers Data Hub (Customer.jsx).

The page used to export by firing one GET per selected row, and "select all"
only ever selected the 50 rows of the current page - so the biggest export the
UI could produce was 50 records, at the cost of 50 round trips.

This module replaces that with ONE generic endpoint that exports every row the
grid is currently showing:

  * rows are streamed out of SQL Server with yield_per(), so memory stays flat
    whatever the table size (asset_detailed is ~41k rows);
  * the page's search term is honoured, so "export all" means "export
    everything the current search matches", not blindly the whole table;
  * the dynamic extra_data JSON is expanded into real columns, exactly the way
    the grid expands it;
  * the workbook is written with openpyxl's write-only sheet, which never keeps
    more than one row of cells alive, straight into a temp file that is then
    streamed to the browser and deleted.

The search columns per table mirror the *_search_filter / *_count methods in
CustomerController - keep the two in sync when a searchable column is added.
"""

import csv
import json
import os
import pickle
import tempfile
from datetime import date, datetime
from typing import Any, Dict, Iterable, List, Optional

from fastapi import APIRouter, Depends, Header, HTTPException, Query, status
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import case, desc, or_, select
from sqlalchemy.orm import Session
from starlette.background import BackgroundTask

from app.database import SessionLocal
from app.controllers.user_controller import UserController
from app.models.customer_model import (
    Customer, AMCAgreement, AssetDetailed, AssetService,
    AnubandhanPlusQuote, AnubandhanQuote, BandhanPlusQuote,
    PulseQuotation, RegularBandhan, LMSData, OpenSRLoadReport,
    MaxTTROilChangeSRZeroLabourFlag, ResponseTimeMaxTTR, CDIDetailReport,
    EFSRReport, AMCExpiryPlanner, LMSInsia, AllInvoiceReport,
)

router = APIRouter(prefix="/customers", tags=["customers"])

# Rows pulled per DBAPI fetch. The remote SQL Server link is latency-bound, so
# a fat batch matters far more than a small buffer: measured on asset_detailed
# (41k rows) a 1000-row batch took ~60s where a 20k-row batch took ~15s.
FETCH_BATCH = 20000

# Columns the grid hides, so the export hides them too.
SYSTEM_FIELDS = ["id", "created_at", "updated_at"]

BOLD = Font(bold=True)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


class ExportTable:
    """One exportable grid tab: which model, how it sorts, what search hits."""

    def __init__(self, model, sheet_name, search, order, hidden=None):
        self.model = model
        self.sheet_name = sheet_name
        self.hidden = set(hidden or SYSTEM_FIELDS)
        # Resolved eagerly so a renamed column blows up at import time, not on
        # the user's first export.
        self.search = [getattr(model, name) for name in search]
        self.order = order(model)

    def filtered(self, db: Session, search: Optional[str], ids: Optional[List[int]]):
        q = db.query(self.model)
        if search and search.strip():
            term = f"%{search.strip()}%"
            q = q.filter(or_(*[col.ilike(term) for col in self.search]))
        if ids:
            q = q.filter(self.model.id.in_(ids))
        return q

    def columns(self) -> List[str]:
        """Fixed columns in declaration order - the grid's own column order."""
        return [
            c.key for c in self.model.__table__.columns
            if c.key not in self.hidden and c.key != "extra_data"
        ]

    @property
    def has_extra(self) -> bool:
        return "extra_data" in self.model.__table__.columns


def _desc(*names):
    return lambda m: [desc(getattr(m, n)) for n in names]


EXPORT_TABLES: Dict[str, ExportTable] = {
    "customers": ExportTable(
        Customer, "Customers",
        ["customer_name", "instance_id", "phone_number", "email", "location", "branch_id"],
        _desc("created_at"),
        hidden=SYSTEM_FIELDS + ["last_updated_by"],
    ),
    "amc_agreements": ExportTable(
        AMCAgreement, "AMC Agreements",
        ["agreement_number", "agreement_name", "instance_id"],
        _desc("agreement_start_date"),
    ),
    "asset_detailed": ExportTable(
        AssetDetailed, "Asset Detailed",
        ["asset_number", "engine_serial_no", "customer_name", "instance_id",
         "krm_number", "krm_status"],
        _desc("created_at"),
    ),
    "asset_services": ExportTable(
        AssetService, "Asset Services",
        ["asset_number", "engine_serial_no", "last_closed_sr_number", "instance_id"],
        _desc("last_oil_change_date"),
    ),
    "anubandhan_plus": ExportTable(
        AnubandhanPlusQuote, "Anubandhan Plus",
        ["quotation_ref_no", "company_name", "engine_no", "instance_id"],
        _desc("created_date_time"),
    ),
    "anubandhan": ExportTable(
        AnubandhanQuote, "Anubandhan",
        ["quotation_ref_no", "company_name", "engine_no", "instance_id"],
        _desc("created_date_time"),
    ),
    "bandhan_plus": ExportTable(
        BandhanPlusQuote, "Bandhan Plus",
        ["quotation_ref_no", "company_name", "engine_no", "instance_id"],
        _desc("created_date_time"),
    ),
    "pulse": ExportTable(
        PulseQuotation, "Pulse Quotations",
        ["quote_id", "account", "instance_id"],
        _desc("creation_date"),
    ),
    "regular_bandhan": ExportTable(
        RegularBandhan, "Regular Bandhan",
        ["quotation_ref_no", "company_name", "engine_no", "instance_id"],
        _desc("created_at"),
    ),
    "lms_data": ExportTable(
        LMSData, "LMS Data",
        ["lead_number", "account_name", "lead_raised_by", "instance_id"],
        _desc("lead_created_date"),
    ),
    "open_sr_load_reports": ExportTable(
        OpenSRLoadReport, "Open SR Load Reports",
        ["service_request_no", "customer_name", "engine_serial_no", "instance_id", "account"],
        _desc("sr_due_date"),
    ),
    "open_sr_data": ExportTable(
        MaxTTROilChangeSRZeroLabourFlag, "MaxTTR Oil Change SR",
        ["instance_id", "branch_id", "branch_name", "account_name", "sr_number",
         "sr_type", "engine_serial_no", "oil_change_flag", "zero_labour_flag"],
        _desc("updated_at"),
    ),
    "response_time_maxttr": ExportTable(
        ResponseTimeMaxTTR, "Response Time MaxTTR",
        ["instance_id", "branch_id", "branch_name", "account_name", "sr_number",
         "sr_type", "engine_serial_no", "se_name", "se_ticket_num"],
        _desc("updated_at"),
    ),
    "amc_expiry_planner": ExportTable(
        AMCExpiryPlanner, "AMC Expiry Planner",
        ["instance_id", "agreement_number", "branch_id", "account_name",
         "installation_site_address"],
        # Soonest expiry first, undated rows last - same order as the grid. The
        # nulls-last flag has to be a CASE: SQL Server has no boolean type, so
        # ordering directly on `col.is_(None)` is a syntax error.
        lambda m: [
            case((m.agreement_end_date.is_(None), 1), else_=0),
            m.agreement_end_date.asc(),
        ],
    ),
    "cdi_detail_report": ExportTable(
        CDIDetailReport, "CDI Detail Report",
        ["instance_id", "sr_number", "branch_name", "x_technician_id",
         "x_technician_name", "cdi_category", "x_account_name", "overall_experience"],
        _desc("updated_at"),
    ),
    "lms_insia": ExportTable(
        LMSInsia, "LMS Data from Insia",
        ["lead_number", "instance_id", "branch_id", "account_name",
         "lead_sr_number", "service_engineer_name"],
        _desc("updated_at"),
    ),
    "efsr_report": ExportTable(
        EFSRReport, "EFSR Report",
        ["instance_id", "service_request_no", "appointment_number", "sd_branch_code",
         "sr_type", "sr_status", "service_engineer_name", "service_engineer_uid", "account"],
        _desc("updated_at"),
    ),
    "all_invoice_report": ExportTable(
        AllInvoiceReport, "All Invoice Detailed Report",
        ["invoice_number", "instance_id", "branch_id", "branch_name", "account_name",
         "invoice_type", "invoice_segment", "invoice_status"],
        # Newest invoice first, undated rows last - same order as the grid.
        lambda m: [
            case((m.invoice_date.is_(None), 1), else_=0),
            m.invoice_date.desc(),
        ],
    ),
}


def _check_export_permission(user_id: str, db: Session):
    user = UserController.get_user_by_id(db, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.role != "admin" and not user.can_export:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to export data",
        )
    return user


def _id_filter(model, ids: List[int]):
    """
    SQL Server caps a statement at 2100 bind parameters, so a big selection has
    to go in as several ORed IN lists rather than one.
    """
    chunks = [ids[i:i + 1000] for i in range(0, len(ids), 1000)]
    return or_(*[model.id.in_(chunk) for chunk in chunks])


def _cell(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (datetime, date, int, float, str)):
        return value
    return str(value)


def _spool(spec: ExportTable, db: Session, search: Optional[str],
           ids: Optional[List[int]], fixed: List[str], spool_path: str):
    """
    Read every matching row ONCE and park it on local disk.

    The header of an xlsx has to be written before any data row, but the set of
    dynamic extra_data columns is only known after every row has been seen. The
    obvious fix - one query for the keys, a second for the rows - pays the
    network cost twice, and against a remote SQL Server that network pass is by
    far the most expensive part of an export (tens of seconds for 41k wide
    rows). So the rows are pickled to a temp file as they arrive and replayed
    from local disk, which is orders of magnitude cheaper than a second pull.

    Returns (ordered extra_data keys, row count).
    """
    columns = [getattr(spec.model, name) for name in fixed]
    if spec.has_extra:
        columns.append(spec.model.extra_data)

    stmt = select(*columns)
    if search and search.strip():
        term = f"%{search.strip()}%"
        stmt = stmt.where(or_(*[col.ilike(term) for col in spec.search]))
    if ids:
        stmt = stmt.where(_id_filter(spec.model, ids))
    stmt = stmt.order_by(*spec.order)
    # Big fetch batches are the single biggest lever on a remote DB: the default
    # trickle of rows turned a 41k-row read from ~15s into over a minute.
    stmt = stmt.execution_options(stream_results=True, max_row_buffer=FETCH_BATCH)

    keys: List[str] = []
    seen = set()
    count = 0
    width = len(fixed)

    with open(spool_path, "wb") as spool:
        dump = pickle.dump
        for row in db.execute(stmt).yield_per(FETCH_BATCH):
            values = [_cell(v) for v in row[:width]]
            extras = None
            if spec.has_extra:
                raw = row[width]
                if raw:
                    try:
                        parsed = json.loads(raw)
                    except (ValueError, TypeError):
                        parsed = None
                    if isinstance(parsed, dict):
                        extras = parsed
                        for key in parsed:
                            if key not in seen:
                                seen.add(key)
                                keys.append(key)
            dump((values, extras), spool, protocol=pickle.HIGHEST_PROTOCOL)
            count += 1

    return keys, count


def _replay(spool_path: str, extra: List[str]) -> Iterable[List[Any]]:
    """Read the spool back and emit finished rows, one at a time."""
    with open(spool_path, "rb") as spool:
        load = pickle.load
        index = 0
        while True:
            try:
                values, extras = load(spool)
            except EOFError:
                return
            index += 1
            row = [index]
            row.extend(values)
            if extra:
                if extras:
                    row.extend(_cell(extras.get(key)) for key in extra)
                else:
                    row.extend([""] * len(extra))
            yield row


def _write_xlsx(path: str, sheet_name: str, header: List[str],
                rows: Iterable[List[Any]]) -> int:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_name[:31])

    # Widths come from the header alone: auto-fitting to the data would mean
    # buffering every row, which is exactly what write-only mode avoids.
    for i, name in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(str(name)) + 4, 12), 45)

    head = []
    for name in header:
        cell = WriteOnlyCell(ws, value=name)
        cell.font = BOLD
        head.append(cell)
    ws.append(head)
    ws.freeze_panes = "A2"

    count = 0
    for row in rows:
        out = []
        for value in row:
            # Real date cells, not strings: Excel's date filters only work on
            # genuine serial dates.
            if isinstance(value, datetime):
                cell = WriteOnlyCell(ws, value=value)
                cell.number_format = (
                    "DD-MM-YYYY"
                    if (value.hour, value.minute, value.second) == (0, 0, 0)
                    else "DD-MM-YYYY HH:MM"
                )
                out.append(cell)
            elif isinstance(value, date):
                cell = WriteOnlyCell(ws, value=value)
                cell.number_format = "DD-MM-YYYY"
                out.append(cell)
            else:
                out.append(value)
        ws.append(out)
        count += 1

    wb.save(path)
    wb.close()
    return count


def _csv_value(v):
    # Match the xlsx side: a midnight timestamp is a plain date, not 00:00.
    if isinstance(v, datetime):
        return (v.strftime("%d-%m-%Y")
                if (v.hour, v.minute, v.second) == (0, 0, 0)
                else v.strftime("%d-%m-%Y %H:%M"))
    if isinstance(v, date):
        return v.strftime("%d-%m-%Y")
    return v


def _write_csv(path: str, header: List[str], rows: Iterable[List[Any]]) -> int:
    count = 0
    # utf-8-sig so Excel picks up the encoding when the CSV is double-clicked.
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(header)
        for row in rows:
            writer.writerow([_csv_value(v) for v in row])
            count += 1
    return count


@router.get("/bulk-export/count")
def bulk_export_count(
    table: str = Query(...),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """How many rows an 'export all' would produce - drives the UI's confirm text."""
    spec = EXPORT_TABLES.get(table)
    if not spec:
        raise HTTPException(status_code=400, detail=f"Unknown table '{table}'")
    return {"table": table, "count": spec.filtered(db, search, None).count()}


@router.get("/bulk-export")
def bulk_export(
    table: str = Query(..., description="Grid tab id, e.g. customers / asset_detailed"),
    search: Optional[str] = Query(None, description="Same search term the grid is filtered by"),
    ids: Optional[str] = Query(None, description="Comma-separated row ids; omit to export everything the search matches"),
    fmt: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    user_id_header: Optional[str] = Header(None, alias="user-id"),
    user_id: Optional[str] = Query(None, description="Alternative to the user-id header, so the browser can download natively"),
    db: Session = Depends(get_db),
):
    """Stream every row of one Customers grid tab as a single Excel / CSV file."""
    resolved_user = user_id_header or user_id
    if not resolved_user:
        raise HTTPException(status_code=401, detail="Missing user id")
    _check_export_permission(resolved_user, db)

    spec = EXPORT_TABLES.get(table)
    if not spec:
        raise HTTPException(status_code=400, detail=f"Unknown table '{table}'")

    id_list: Optional[List[int]] = None
    if ids:
        try:
            id_list = [int(part) for part in ids.split(",") if part.strip()]
        except ValueError:
            raise HTTPException(status_code=400, detail="ids must be comma-separated integers")
        if not id_list:
            id_list = None

    fixed = spec.columns()

    suffix = ".xlsx" if fmt == "xlsx" else ".csv"
    handle, path = tempfile.mkstemp(prefix=f"{table}_export_", suffix=suffix)
    os.close(handle)
    handle, spool_path = tempfile.mkstemp(prefix=f"{table}_spool_", suffix=".pkl")
    os.close(handle)

    try:
        # One trip to the database, then everything else off local disk.
        extra, _count = _spool(spec, db, search, id_list, fixed, spool_path)
        header = ["Sr. No."] + fixed + extra
        rows = _replay(spool_path, extra)
        if fmt == "xlsx":
            _write_xlsx(path, spec.sheet_name, header, rows)
        else:
            _write_csv(path, header, rows)
    except Exception:
        if os.path.exists(path):
            os.unlink(path)
        raise
    finally:
        if os.path.exists(spool_path):
            os.unlink(spool_path)

    stamp = datetime.now().strftime("%Y-%m-%d")
    scope = "selected" if id_list else ("filtered" if (search or "").strip() else "all")
    filename = f"{table}_{scope}_{stamp}{suffix}"
    media = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if fmt == "xlsx" else "text/csv"
    )

    def _cleanup():
        if os.path.exists(path):
            os.unlink(path)

    return FileResponse(
        path,
        media_type=media,
        filename=filename,
        background=BackgroundTask(_cleanup),
    )
