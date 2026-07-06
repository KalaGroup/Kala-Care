from sqlalchemy.orm import Session
from sqlalchemy import func, desc, and_, or_, case, Integer, Numeric, String, Date, between, Time, cast, extract, text
from datetime import datetime, timedelta, timezone
from typing import Optional, List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

from app.models.engagement_model import FollowUp, Activity, RR
from app.models.user_model import User
from app.models.campaign_model import Campaign

# IST timezone (UTC+5:30) — created_at is stored in IST via SQL Server GETDATE()
IST = timezone(timedelta(hours=5, minutes=30))

# Branch name mapping helper
BRANCH_NAME_MAP = {
    '420435_1': 'Ch.Sambhaji Nagar',
    '420435_2': 'Ahilyanagar',
    '420435_3': 'Beed',
    '420435_4': 'Nanded',
    '420435_5': 'Babhaleshwar',
    '420435_6': 'Latur',
    '420435_7': 'Parbhani',
    '420435_8': 'Hubli',
    '420435_9': 'Belagavi',
    '420435_10': 'Hospet',
    '420435_11': 'Ballari',
    '420435_12': 'Bagalkot',
    '420435_13': 'Gulbarga',
    '420435_14': 'Bijapur'
}

def get_branch_display_name(branch_code):
    """Helper function to get branch display name"""
    if not branch_code:
        return 'N/A'
    branch_name = BRANCH_NAME_MAP.get(branch_code)
    if branch_name:
        return f"{branch_name} ({branch_code})"
    return branch_code

class EmployeePerformanceController:
    
    @staticmethod
    def apply_time_filter(query, model, time_period: str = 'all', start_date=None, end_date=None):
        """
        Apply time filter to query based on time_period or custom date range.
        created_at is stored in IST (SQL Server GETDATE()), so ALL boundaries
        must be IST-naive datetimes — never UTC.
        """
        # Custom range: start_date/end_date arrive as IST-naive datetimes from the route
        if start_date is not None and end_date is not None:
            if isinstance(start_date, datetime) and isinstance(end_date, datetime):
                return query.filter(model.created_at >= start_date, model.created_at <= end_date)
            elif isinstance(start_date, str) and isinstance(end_date, str):
                try:
                    start = datetime.strptime(start_date, '%Y-%m-%d')
                    # Add one day to end date to include the full end date
                    end = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                    return query.filter(model.created_at >= start, model.created_at <= end)
                except Exception as e:
                    print(f"Error parsing custom dates: {e}")
                    return query
            else:
                return query

        if time_period and time_period != 'all':
            # IST "now" as a naive datetime to match IST-stored created_at
            now = datetime.now(IST).replace(tzinfo=None)
            if time_period == 'month':
                cutoff_date = now - timedelta(days=30)
            elif time_period == '3months':
                cutoff_date = now - timedelta(days=90)
            elif time_period == '6months':
                cutoff_date = now - timedelta(days=180)
            elif time_period == 'year':
                cutoff_date = now - timedelta(days=365)
            else:
                return query

            return query.filter(model.created_at >= cutoff_date)

        return query
    
    @staticmethod
    def get_daily_performance_with_details(db: Session, user_id: str, time_period: str = 'all', start_date: str = None, end_date: str = None):
        """
        Get daily performance breakdown including working hours and follow-up types with status breakdown
        Includes rescheduled status
        """
        try:            
            # Build date filter
            date_condition = ""
            params = {'user_id': user_id}
            
            if time_period == 'custom' and start_date and end_date:
                date_condition = "AND created_at >= :start_date AND created_at <= :end_date"
                params['start_date'] = start_date
                params['end_date'] = end_date
            elif time_period != 'all' and time_period != 'custom':
                if time_period == 'month':
                    date_condition = "AND created_at >= DATEADD(day, -30, GETUTCDATE())"
                elif time_period == '3months':
                    date_condition = "AND created_at >= DATEADD(day, -90, GETUTCDATE())"
                elif time_period == '6months':
                    date_condition = "AND created_at >= DATEADD(day, -180, GETUTCDATE())"
                elif time_period == 'year':
                    date_condition = "AND created_at >= DATEADD(day, -365, GETUTCDATE())"
            
            # SQL query now COMBINES followups + non_followups for the same user.
            # date_condition uses an unqualified "created_at", so it applies correctly
            # inside each UNION branch (one table per branch).
            sql_query = f"""
                WITH combined_followups AS (
                    SELECT f.created_at, f.followup_by, f.status, f.campaign_id
                    FROM followups f
                    WHERE f.user_id = :user_id
                        AND f.created_at IS NOT NULL
                        {date_condition}
                    UNION ALL
                    SELECT n.created_at, n.followup_by, n.status, n.campaign_id
                    FROM non_followups n
                    WHERE n.user_id = :user_id
                        AND n.created_at IS NOT NULL
                        {date_condition}
                )
                SELECT 
                    CONVERT(DATE, cf.created_at) as date,
                    MIN(cf.created_at) as first_followup_time,
                    MAX(cf.created_at) as last_followup_time,
                    COUNT(*) as total_followups,
                    -- Campaign NAME (not id) resolved from campaign_id across BOTH tables for this day
                    STUFF((
                        SELECT DISTINCT ', ' + c2.name
                        FROM campaigns c2
                        WHERE c2.id IN (
                            SELECT DISTINCT cf2.campaign_id
                            FROM combined_followups cf2
                            WHERE CONVERT(DATE, cf2.created_at) = CONVERT(DATE, cf.created_at)
                                AND cf2.campaign_id IS NOT NULL
                        )
                        FOR XML PATH('')
                    ), 1, 2, '') as campaign_name,
                    SUM(CASE WHEN cf.followup_by = 'call' THEN 1 ELSE 0 END) as followup_by_call,
                    SUM(CASE WHEN cf.followup_by = 'call' AND cf.status = 'completed' THEN 1 ELSE 0 END) as call_completed,
                    SUM(CASE WHEN cf.followup_by = 'call' AND cf.status = 'wip' THEN 1 ELSE 0 END) as call_wip,
                    SUM(CASE WHEN cf.followup_by = 'call' AND cf.status = 'rejected' THEN 1 ELSE 0 END) as call_rejected,
                    SUM(CASE WHEN cf.followup_by = 'call' AND cf.status = 'rescheduled' THEN 1 ELSE 0 END) as call_rescheduled,
                    -- WhatsApp breakdown
                    SUM(CASE WHEN cf.followup_by = 'whatsapp' THEN 1 ELSE 0 END) as followup_by_whatsapp,
                    SUM(CASE WHEN cf.followup_by = 'whatsapp' AND cf.status = 'completed' THEN 1 ELSE 0 END) as whatsapp_completed,
                    SUM(CASE WHEN cf.followup_by = 'whatsapp' AND cf.status = 'wip' THEN 1 ELSE 0 END) as whatsapp_wip,
                    SUM(CASE WHEN cf.followup_by = 'whatsapp' AND cf.status = 'rejected' THEN 1 ELSE 0 END) as whatsapp_rejected,
                    SUM(CASE WHEN cf.followup_by = 'whatsapp' AND cf.status = 'rescheduled' THEN 1 ELSE 0 END) as whatsapp_rescheduled,
                    -- Email breakdown
                    SUM(CASE WHEN cf.followup_by = 'email' THEN 1 ELSE 0 END) as followup_by_email,
                    SUM(CASE WHEN cf.followup_by = 'email' AND cf.status = 'completed' THEN 1 ELSE 0 END) as email_completed,
                    SUM(CASE WHEN cf.followup_by = 'email' AND cf.status = 'wip' THEN 1 ELSE 0 END) as email_wip,
                    SUM(CASE WHEN cf.followup_by = 'email' AND cf.status = 'rejected' THEN 1 ELSE 0 END) as email_rejected,
                    SUM(CASE WHEN cf.followup_by = 'email' AND cf.status = 'rescheduled' THEN 1 ELSE 0 END) as email_rescheduled,
                    -- Visit breakdown
                    SUM(CASE WHEN cf.followup_by = 'visit' THEN 1 ELSE 0 END) as followup_by_visit,
                    SUM(CASE WHEN cf.followup_by = 'visit' AND cf.status = 'completed' THEN 1 ELSE 0 END) as visit_completed,
                    SUM(CASE WHEN cf.followup_by = 'visit' AND cf.status = 'wip' THEN 1 ELSE 0 END) as visit_wip,
                    SUM(CASE WHEN cf.followup_by = 'visit' AND cf.status = 'rejected' THEN 1 ELSE 0 END) as visit_rejected,
                    SUM(CASE WHEN cf.followup_by = 'visit' AND cf.status = 'rescheduled' THEN 1 ELSE 0 END) as visit_rescheduled,
                    -- NC breakdown (all 4 together at the end so row[25]-row[28] match Python)
                    SUM(CASE WHEN cf.followup_by = 'call'     AND cf.status = 'not_connected' THEN 1 ELSE 0 END) as call_not_connected,
                    SUM(CASE WHEN cf.followup_by = 'whatsapp' AND cf.status = 'not_connected' THEN 1 ELSE 0 END) as whatsapp_not_connected,
                    SUM(CASE WHEN cf.followup_by = 'email'    AND cf.status = 'not_connected' THEN 1 ELSE 0 END) as email_not_connected,
                    SUM(CASE WHEN cf.followup_by = 'visit'    AND cf.status = 'not_connected' THEN 1 ELSE 0 END) as visit_not_connected
                FROM combined_followups cf
                GROUP BY CONVERT(DATE, cf.created_at)
                ORDER BY CONVERT(DATE, cf.created_at) DESC
            """
            
            result = db.execute(text(sql_query), params)
            rows = result.fetchall()
            
            daily_performance = []
            for row in rows:
                date_val = row[0]
                first_time = row[1]
                last_time = row[2]
                total_followups = int(row[3]) if row[3] else 0
                campaign_name = row[4] if row[4] else 'N/A'
                
                # Calculate working hours
                working_hours = None
                if first_time and last_time:
                    diff = last_time - first_time
                    working_hours = round(diff.total_seconds() / 3600, 2)
                
                daily_performance.append({
                    'date': date_val.strftime('%Y-%m-%d') if hasattr(date_val, 'strftime') else str(date_val),
                    'campaign_name': campaign_name,
                    'first_followup_time': first_time.isoformat() if first_time else None,
                    'last_followup_time': last_time.isoformat() if last_time else None,
                    'total_working_hours': working_hours,
                    'total_followups': total_followups,
                    'completed_count': 0,
                    # Call
                    'followup_by_call': int(row[5]) if row[5] else 0,
                    'call_completed': int(row[6]) if row[6] else 0,
                    'call_wip': int(row[7]) if row[7] else 0,
                    'call_rejected': int(row[8]) if row[8] else 0,
                    'call_rescheduled': int(row[9]) if row[9] else 0,
                    # WhatsApp
                    'followup_by_whatsapp': int(row[10]) if row[10] else 0,
                    'whatsapp_completed': int(row[11]) if row[11] else 0,
                    'whatsapp_wip': int(row[12]) if row[12] else 0,
                    'whatsapp_rejected': int(row[13]) if row[13] else 0,
                    'whatsapp_rescheduled': int(row[14]) if row[14] else 0,
                    # Email
                    'followup_by_email': int(row[15]) if row[15] else 0,
                    'email_completed': int(row[16]) if row[16] else 0,
                    'email_wip': int(row[17]) if row[17] else 0,
                    'email_rejected': int(row[18]) if row[18] else 0,
                    'email_rescheduled': int(row[19]) if row[19] else 0,
                    # Visit
                    'followup_by_visit': int(row[20]) if row[20] else 0,
                    'visit_completed': int(row[21]) if row[21] else 0,
                    'visit_wip': int(row[22]) if row[22] else 0,
                    'visit_rejected': int(row[23]) if row[23] else 0,
                    'visit_rescheduled': int(row[24]) if row[24] else 0,
                    'call_not_connected': int(row[25]) if row[25] else 0,
                    'whatsapp_not_connected': int(row[26]) if row[26] else 0,
                    'email_not_connected': int(row[27]) if row[27] else 0,
                    'visit_not_connected': int(row[28]) if row[28] else 0
                })
                            
            return daily_performance
            
        except Exception as e:
            print(f"Error in get_daily_performance_with_details: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    @staticmethod
    async def get_employee_performance(db: Session, user_id: str, user_name: str = None, branch: str = None, time_period: str = 'all', start_date=None, end_date=None):
        """
        Get real-time performance for a specific employee with time filter
        Includes rescheduled status and quotation only for completed follow-ups
        """
        try:            
            user_info = db.query(User).filter(User.user_id == user_id).first()
            
            # Base query
            base_query = db.query(FollowUp).filter(FollowUp.user_id == user_id)
            
            # Apply time filter
            filtered_query = EmployeePerformanceController.apply_time_filter(base_query, FollowUp, time_period, start_date, end_date)
            
            # Get follow-up counts by status (including rescheduled)
            # Quotation only counted when status is 'completed'
            status_counts = filtered_query.with_entities(
                func.count(FollowUp.id).label('total'),
                func.sum(case((FollowUp.status == 'completed', 1), else_=0)).label('completed'),
                func.sum(case((FollowUp.status == 'wip', 1), else_=0)).label('wip'),
                func.sum(case((FollowUp.status == 'rejected', 1), else_=0)).label('rejected'),
                func.sum(case((FollowUp.status == 'pending', 1), else_=0)).label('pending'),
                func.sum(case((FollowUp.status == 'rescheduled', 1), else_=0)).label('rescheduled'),
                # Only count quotation when status is 'completed'
                func.sum(case((and_(FollowUp.quotation_sent == True, FollowUp.status == 'completed'), FollowUp.quotation_value), else_=0)).label('total_quotation'),
                func.sum(case((and_(FollowUp.quotation_sent == True, FollowUp.status == 'completed'), 1), else_=0)).label('quotations_sent')
            ).first()
            
            # Get follow-up type breakdown (followup_by - for call/whatsapp/email/visit)
            type_breakdown_raw = filtered_query.with_entities(
                FollowUp.followup_by,
                func.count(FollowUp.id).label('count')
            ).filter(
                FollowUp.followup_by.isnot(None)
            ).group_by(FollowUp.followup_by).all()
            
            type_breakdown = {}
            for item in type_breakdown_raw:
                if item.followup_by:
                    type_breakdown[item.followup_by] = int(item.count)
            
            # Get follow-up flag breakdown (followup_flag - for C1-C7)
            flag_breakdown_raw = filtered_query.with_entities(
                FollowUp.followup_flag,
                func.count(FollowUp.id).label('count')
            ).filter(
                FollowUp.followup_flag.isnot(None)
            ).group_by(FollowUp.followup_flag).all()
            
            flag_breakdown = {}
            for item in flag_breakdown_raw:
                if item.followup_flag:
                    flag_breakdown[item.followup_flag] = int(item.count)
            
            # Get daily stats
            daily_stats_query = db.query(
                func.cast(FollowUp.created_at, Date).label('date'),
                func.count(FollowUp.id).label('total'),
                func.sum(case((FollowUp.status == 'completed', 1), else_=0)).label('completed')
            ).filter(
                FollowUp.user_id == user_id
            )
            
            daily_stats_query = EmployeePerformanceController.apply_time_filter(daily_stats_query, FollowUp, time_period, start_date, end_date)
            
            daily_stats_raw = daily_stats_query.group_by(
                func.cast(FollowUp.created_at, Date)
            ).order_by(
                func.cast(FollowUp.created_at, Date).desc()
            ).all()
            
            daily_stats = {}
            for item in daily_stats_raw:
                if item.date:
                    date_str = item.date.strftime('%Y-%m-%d')
                    daily_stats[date_str] = {
                        'total': int(item.total) if item.total else 0,
                        'completed': int(item.completed) if item.completed else 0
                    }
            
            # Get detailed daily performance
            daily_performance = EmployeePerformanceController.get_daily_performance_with_details(
                db, user_id, time_period, start_date, end_date
            )
                        
            performance = {
                'user_id': user_id,
                'user_name': user_name or (user_info.name if user_info else ''),
                'branch': branch or (user_info.branch if user_info else None),
                'total_followups': int(status_counts.total) if status_counts.total else 0,
                'completed_count': int(status_counts.completed) if status_counts.completed else 0,
                'wip_count': int(status_counts.wip) if status_counts.wip else 0,
                'rejected_count': int(status_counts.rejected) if status_counts.rejected else 0,
                'rescheduled_count': int(status_counts.rescheduled) if status_counts.rescheduled else 0,
                'pending_count': int(status_counts.pending) if status_counts.pending else 0,
                'total_quotation_value': float(status_counts.total_quotation) if status_counts.total_quotation else 0,
                'quotations_sent_count': int(status_counts.quotations_sent) if status_counts.quotations_sent else 0,
                'followup_type_breakdown': type_breakdown,
                'followup_flag_breakdown': flag_breakdown,
                'daily_stats': daily_stats,
                'daily_performance': daily_performance,
                'can_export': user_info.can_export if user_info else False,
                'time_period': time_period,
                'start_date': start_date,
                'end_date': end_date
            }
            
            return performance
            
        except Exception as e:
            print(f"Error in get_employee_performance: {str(e)}")
            import traceback
            traceback.print_exc()
            raise e

    @staticmethod
    async def get_all_employees_performance(db: Session, branch_code: str = None, time_period: str = 'all', start_date=None, end_date=None):
        """
        Get real-time performance for all employees with time filter.
        OPTIMIZED: single grouped query instead of N+1 (one per user).
        Heavy fields (daily_stats, daily_performance, followup_type_breakdown, followup_flag_breakdown)
        are returned empty — same pattern as get_employees_by_branch.
        """
        try:
            user_query = db.query(User).filter(User.is_blocked == False)
            if branch_code:
                user_query = user_query.filter(User.branch == branch_code)
            
            users = user_query.all()
            if not users:
                return []
            
            user_ids = [u.user_id for u in users]
            
            # ONE grouped query for all users instead of N+ queries per user
            base_query = db.query(FollowUp).filter(FollowUp.user_id.in_(user_ids))
            filtered_query = EmployeePerformanceController.apply_time_filter(
                base_query, FollowUp, time_period, start_date, end_date
            )
            
            per_user_stats = filtered_query.with_entities(
                FollowUp.user_id,
                func.count(FollowUp.id).label('total'),
                func.sum(case((FollowUp.status == 'completed', 1), else_=0)).label('completed'),
                func.sum(case((FollowUp.status == 'wip', 1), else_=0)).label('wip'),
                func.sum(case((FollowUp.status == 'rejected', 1), else_=0)).label('rejected'),
                func.sum(case((FollowUp.status == 'rescheduled', 1), else_=0)).label('rescheduled'),
                func.sum(case((FollowUp.status == 'pending', 1), else_=0)).label('pending'),
                # "Not Connected" is now its own status value.
                func.sum(case((FollowUp.status == 'not_connected', 1), else_=0)).label('not_connected'),
                func.sum(case((and_(FollowUp.quotation_sent == True, FollowUp.status == 'completed'), FollowUp.quotation_value), else_=0)).label('total_quotation'),
                func.sum(case((and_(FollowUp.quotation_sent == True, FollowUp.status == 'completed'), 1), else_=0)).label('quotations_sent'),
            ).group_by(FollowUp.user_id).all()
            
            stats_map = {row.user_id: row for row in per_user_stats}
            
            performances = []
            for user in users:
                stats = stats_map.get(user.user_id)
                performances.append({
                    'user_id': user.user_id,
                    'user_name': user.name,
                    'branch': user.branch,
                    'total_followups': int(stats.total or 0) if stats else 0,
                    'completed_count': int(stats.completed or 0) if stats else 0,
                    'wip_count': int(stats.wip or 0) if stats else 0,
                    'rejected_count': int(stats.rejected or 0) if stats else 0,
                    'rescheduled_count': int(stats.rescheduled or 0) if stats else 0,
                    'pending_count': int(stats.pending or 0) if stats else 0,
                    'not_connected_count': int(stats.not_connected or 0) if stats else 0,
                    'total_quotation_value': float(stats.total_quotation or 0) if stats else 0,
                    'quotations_sent_count': int(stats.quotations_sent or 0) if stats else 0,
                    'followup_type_breakdown': {},
                    'followup_flag_breakdown': {},
                    'daily_stats': {},
                    'daily_performance': [],
                    'can_export': user.can_export if hasattr(user, 'can_export') else False,
                    'time_period': time_period,
                    'start_date': start_date,
                    'end_date': end_date
                })
            
            performances.sort(key=lambda x: x['completed_count'], reverse=True)
            return performances
            
        except Exception as e:
            print(f"Error in get_all_employees_performance: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    @staticmethod
    async def get_branch_performance(db: Session, branch_code: str = None, time_period: str = 'all', start_date=None, end_date=None):
        """
        Get real-time branch-wise performance with time filter
        Includes rescheduled status and quotation only for completed
        """
        try:
            branch_query = db.query(User.branch).filter(User.is_blocked == False).distinct()
            if branch_code:
                branch_query = branch_query.filter(User.branch == branch_code)
            
            branches = [b[0] for b in branch_query.all() if b[0]]
            
            branch_performances = []
            
            for branch in branches:
                users = db.query(User).filter(
                    User.branch == branch,
                    User.is_blocked == False
                ).all()
                
                user_ids = [u.user_id for u in users]
                
                if not user_ids:
                    continue
                
                base_query = db.query(FollowUp).filter(FollowUp.user_id.in_(user_ids))
                filtered_query = EmployeePerformanceController.apply_time_filter(base_query, FollowUp, time_period, start_date, end_date)
                
                stats = filtered_query.with_entities(
                    func.count(FollowUp.id).label('total'),
                    func.sum(case((FollowUp.status == 'completed', 1), else_=0)).label('completed'),
                    func.sum(case((FollowUp.status == 'wip', 1), else_=0)).label('wip'),
                    func.sum(case((FollowUp.status == 'rejected', 1), else_=0)).label('rejected'),
                    func.sum(case((FollowUp.status == 'pending', 1), else_=0)).label('pending'),
                    func.sum(case((FollowUp.status == 'rescheduled', 1), else_=0)).label('rescheduled'),
                    # Only count quotation when status is 'completed'
                    func.sum(case((and_(FollowUp.quotation_sent == True, FollowUp.status == 'completed'), FollowUp.quotation_value), else_=0)).label('total_quotation'),
                    func.sum(case((and_(FollowUp.quotation_sent == True, FollowUp.status == 'completed'), 1), else_=0)).label('quotations_sent')
                ).first()
                
                # Single grouped query instead of N queries (one per user)
                per_user_stats = filtered_query.with_entities(
                    FollowUp.user_id,
                    FollowUp.user_name,
                    func.count(FollowUp.id).label('total'),
                    func.sum(case((FollowUp.status == 'completed', 1), else_=0)).label('completed'),
                ).group_by(FollowUp.user_id, FollowUp.user_name).all()

                user_export_map = {u.user_id: u.can_export for u in users}

                employee_performances = [{
                    'user_id': r.user_id,
                    'user_name': r.user_name,
                    'completed_count': int(r.completed or 0),
                    'total_followups': int(r.total or 0),
                    'can_export': user_export_map.get(r.user_id, False),
                } for r in per_user_stats]

                employee_performances.sort(key=lambda x: x['completed_count'], reverse=True)
                top_performers = [{
                    'user_id': perf['user_id'],
                    'user_name': perf['user_name'],
                    'completed': perf['completed_count'],
                    'total': perf['total_followups'],
                    'can_export': perf['can_export']
                } for perf in employee_performances[:5]]
                
                top_campaigns = await EmployeePerformanceController.get_branch_top_campaigns(
                    db, branch, time_period, start_date, end_date
                )
                
                branch_performances.append({
                    'branch': branch,
                    'total_employees': len(users),
                    'total_followups': int(stats.total) if stats.total else 0,
                    'completed_count': int(stats.completed) if stats.completed else 0,
                    'wip_count': int(stats.wip) if stats.wip else 0,
                    'rejected_count': int(stats.rejected) if stats.rejected else 0,
                    'rescheduled_count': int(stats.rescheduled) if stats.rescheduled else 0,
                    'pending_count': int(stats.pending) if stats.pending else 0,
                    'total_quotation_value': float(stats.total_quotation) if stats.total_quotation else 0,
                    'quotations_sent_count': int(stats.quotations_sent) if stats.quotations_sent else 0,
                    'employee_summary': top_performers,
                    'top_campaigns': top_campaigns,
                    'time_period': time_period,
                    'start_date': start_date,
                    'end_date': end_date
                })
            
            return branch_performances
            
        except Exception as e:
            print(f"Error in get_branch_performance: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    @staticmethod
    async def get_employees_by_branch(db: Session, branch_code: str, time_period: str = 'all', start_date=None, end_date=None):
        """
        Get all employees for a specific branch with their performance data.
        Optimized: single grouped query instead of one per employee.
        """
        try:
            users = db.query(User).filter(
                User.branch == branch_code,
                User.is_blocked == False
            ).all()
            
            if not users:
                return []
            
            user_ids = [u.user_id for u in users]
            
            # ONE grouped query for all employees in this branch
            base_query = db.query(FollowUp).filter(FollowUp.user_id.in_(user_ids))
            filtered_query = EmployeePerformanceController.apply_time_filter(
                base_query, FollowUp, time_period, start_date, end_date
            )
            
            per_user_stats = filtered_query.with_entities(
                FollowUp.user_id,
                func.count(FollowUp.id).label('total'),
                func.sum(case((FollowUp.status == 'completed', 1), else_=0)).label('completed'),
                func.sum(case((FollowUp.status == 'wip', 1), else_=0)).label('wip'),
                func.sum(case((FollowUp.status == 'rejected', 1), else_=0)).label('rejected'),
                func.sum(case((FollowUp.status == 'rescheduled', 1), else_=0)).label('rescheduled'),
                func.sum(case((FollowUp.status == 'pending', 1), else_=0)).label('pending'),
                # "Not Connected" is now its own status value.
                func.sum(case((FollowUp.status == 'not_connected', 1), else_=0)).label('not_connected'),
                func.sum(case((and_(FollowUp.quotation_sent == True, FollowUp.status == 'completed'), FollowUp.quotation_value), else_=0)).label('total_quotation'),
                func.sum(case((and_(FollowUp.quotation_sent == True, FollowUp.status == 'completed'), 1), else_=0)).label('quotations_sent'),
            ).group_by(FollowUp.user_id).all()
            
            # Map by user_id for quick lookup
            stats_map = {row.user_id: row for row in per_user_stats}
            
            employees = []
            for user in users:
                stats = stats_map.get(user.user_id)
                employees.append({
                    'user_id': user.user_id,
                    'user_name': user.name,
                    'branch': user.branch,
                    'total_followups': int(stats.total or 0) if stats else 0,
                    'completed_count': int(stats.completed or 0) if stats else 0,
                    'wip_count': int(stats.wip or 0) if stats else 0,
                    'rejected_count': int(stats.rejected or 0) if stats else 0,
                    'rescheduled_count': int(stats.rescheduled or 0) if stats else 0,
                    'pending_count': int(stats.pending or 0) if stats else 0,
                    'not_connected_count': int(stats.not_connected or 0) if stats else 0,
                    'total_quotation_value': float(stats.total_quotation or 0) if stats else 0,
                    'quotations_sent_count': int(stats.quotations_sent or 0) if stats else 0,
                    'followup_type_breakdown': {},
                    'followup_flag_breakdown': {},
                    'daily_stats': {},
                    'daily_performance': [],
                    'can_export': user.can_export if hasattr(user, 'can_export') else False,
                    'time_period': time_period,
                    'start_date': start_date,
                    'end_date': end_date
                })
            
            employees.sort(key=lambda x: x['completed_count'], reverse=True)
            return employees
            
        except Exception as e:
            print(f"Error in get_employees_by_branch: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    @staticmethod
    async def get_all_branches(db: Session):
        """
        Get all unique branch codes from users table
        """
        try:
            branches = db.query(User.branch).filter(
                User.branch.isnot(None),
                User.is_blocked == False
            ).distinct().all()
            return [branch[0] for branch in branches if branch[0]]
        except Exception as e:
            print(f"Error in get_all_branches: {str(e)}")
            return []

    @staticmethod
    async def get_performance_summary(db: Session, time_period: str = 'all', start_date=None, end_date=None):
        """
        Get overall performance summary with time filter
        """
        try:
            from app.models.campaign_model import Campaign
            from app.models.engagement_model import FollowUp
            from app.models.user_model import User
            from sqlalchemy import and_
            
            # Get all active campaigns
            active_campaigns = db.query(Campaign).filter(Campaign.status == 'active').all()
            active_campaign_ids = [c.id for c in active_campaigns]
            
            # Step 1: Total Assets = asset_numbers array from active campaigns (with duplicates as per array)
            total_assets_list = []
            
            for campaign in active_campaigns:
                asset_numbers = campaign.asset_numbers or []
                for asset in asset_numbers:
                    if asset:
                        total_assets_list.append(asset)
            
            # Also add completed follow-ups from active campaigns - WITH TIME FILTER
            completed_fu_query = db.query(FollowUp).filter(
                FollowUp.campaign_id.in_(active_campaign_ids),
                FollowUp.status == 'completed'
            )
            completed_fu_query = EmployeePerformanceController.apply_time_filter(
                completed_fu_query, FollowUp, time_period, start_date, end_date
            )
            completed_followups = completed_fu_query.all()
            
            for fu in completed_followups:
                if fu.customer_instance_id:
                    total_assets_list.append(fu.customer_instance_id)
            
            # Get additional customers from followups (not in asset_numbers and not completed) - WITH TIME FILTER
            # OPTIMIZED: ONE grouped query for all active campaigns instead of N queries
            if active_campaign_ids:
                followup_only_query = db.query(
                    FollowUp.campaign_id,
                    FollowUp.customer_instance_id
                ).filter(
                    FollowUp.campaign_id.in_(active_campaign_ids),
                    FollowUp.customer_instance_id.isnot(None),
                    FollowUp.status != 'completed'
                ).distinct()
                followup_only_query = EmployeePerformanceController.apply_time_filter(
                    followup_only_query, FollowUp, time_period, start_date, end_date
                )
                followup_only_rows = followup_only_query.all()
                
                # Group by campaign_id in Python
                followup_only_by_campaign = {}
                for cid, iid in followup_only_rows:
                    if iid:
                        followup_only_by_campaign.setdefault(cid, set()).add(iid)
                
                for campaign in active_campaigns:
                    asset_numbers_set = set(campaign.asset_numbers or [])
                    for customer in followup_only_by_campaign.get(campaign.id, set()):
                        if customer not in asset_numbers_set:
                            total_assets_list.append(customer)
            
            total_assets_count = len(total_assets_list)
            
            # Step 2: Get unique assets for customer counting
            unique_assets = set(total_assets_list)
            total_customers_count = len(unique_assets)
            
            # Step 3: Get all followups for active campaigns only - WITH TIME FILTER
            followup_base = db.query(
                FollowUp.customer_instance_id,
                FollowUp.campaign_id,
                FollowUp.status,
                FollowUp.followup_date,
                FollowUp.created_at
            ).filter(
                FollowUp.campaign_id.in_(active_campaign_ids)
            )
            followup_base = EmployeePerformanceController.apply_time_filter(
                followup_base, FollowUp, time_period, start_date, end_date
            )
            all_followups = followup_base.all()
            
            # Step 4: Track unique (instance_id + campaign_id) combinations for assets
            # Key: f"{instance_id}_{campaign_id}"
            asset_combination_tracker = {}

            # Hoisted: set membership is O(1) vs O(n) list scan per follow-up
            active_campaign_ids_set = set(active_campaign_ids)

            for fu in all_followups:
                instance_id = fu.customer_instance_id
                campaign_id = fu.campaign_id

                if instance_id and campaign_id in active_campaign_ids_set:
                    key = f"{instance_id}_{campaign_id}"
                    followup_date = fu.followup_date if fu.followup_date else fu.created_at
                    
                    if key not in asset_combination_tracker:
                        asset_combination_tracker[key] = {
                            'status': fu.status,
                            'date': followup_date,
                            'instance_id': instance_id,
                            'campaign_id': campaign_id
                        }
                    else:
                        # Keep the latest follow-up
                        if followup_date > asset_combination_tracker[key]['date']:
                            asset_combination_tracker[key]['status'] = fu.status
                            asset_combination_tracker[key]['date'] = followup_date
            
            # Step 5: Calculate asset metrics
            # Attended Assets = count of unique (instance_id + campaign_id) combinations
            attended_assets_count = len(asset_combination_tracker)
            
            # Count status breakdown for assets
            completed_assets = 0
            wip_assets = 0
            rejected_assets = 0
            rescheduled_assets = 0
            pending_assets = 0
            not_connected_assets = 0  # NC: latest status = not_connected
            
            for key, data in asset_combination_tracker.items():
                status = data['status']
                if status == 'completed':
                    completed_assets += 1
                elif status == 'wip':
                    wip_assets += 1
                elif status == 'rejected':
                    rejected_assets += 1
                elif status == 'rescheduled':
                    rescheduled_assets += 1
                elif status == 'not_connected':
                    not_connected_assets += 1
                elif status == 'pending':
                    pending_assets += 1
            
            # Verify: attended_assets_count should equal completed + wip + rejected + rescheduled
            remaining_assets_count = total_assets_count - attended_assets_count
            
            # Step 6: Calculate customer metrics (based on unique instance_id from followups with active campaigns)
            attended_customers_set = set()
            for fu in all_followups:
                if fu.customer_instance_id:
                    attended_customers_set.add(fu.customer_instance_id)
            
            attended_customers_count = len(attended_customers_set)
            remaining_customers_count = total_customers_count - attended_customers_count
            
            # Step 7: Get follow-up counts for active campaigns only
            if active_campaign_ids:
                base_query = db.query(FollowUp).filter(
                    FollowUp.campaign_id.in_(active_campaign_ids)
                )
                
                # Apply time filter
                filtered_query = EmployeePerformanceController.apply_time_filter(base_query, FollowUp, time_period, start_date, end_date)
                
                stats = filtered_query.with_entities(
                    func.count(FollowUp.id).label('total'),
                    func.sum(case((FollowUp.status == 'completed', 1), else_=0)).label('completed'),
                    func.sum(case((FollowUp.status == 'wip', 1), else_=0)).label('wip'),
                    func.sum(case((FollowUp.status == 'rejected', 1), else_=0)).label('rejected'),
                    func.sum(case((FollowUp.status == 'pending', 1), else_=0)).label('pending'),
                    func.sum(case((FollowUp.status == 'rescheduled', 1), else_=0)).label('rescheduled'),
                    # Quotation only for completed status in active campaigns
                    func.sum(case((and_(FollowUp.quotation_sent == True, FollowUp.status == 'completed'), FollowUp.quotation_value), else_=0)).label('total_quotation'),
                    func.sum(case((and_(FollowUp.quotation_sent == True, FollowUp.status == 'completed'), 1), else_=0)).label('quotations_sent')
                ).first()
                
                total_followups = int(stats.total) if stats.total else 0
                completed_followups = int(stats.completed) if stats.completed else 0
                wip_followups = int(stats.wip) if stats.wip else 0
                rejected_followups = int(stats.rejected) if stats.rejected else 0
                rescheduled_followups = int(stats.rescheduled) if stats.rescheduled else 0
                pending_followups = int(stats.pending) if stats.pending else 0
                total_quotation = float(stats.total_quotation) if stats.total_quotation else 0
                quotations_sent = int(stats.quotations_sent) if stats.quotations_sent else 0
            else:
                total_followups = 0
                completed_followups = 0
                wip_followups = 0
                rejected_followups = 0
                rescheduled_followups = 0
                pending_followups = 0
                total_quotation = 0
                quotations_sent = 0
            
            return {
                # Asset metrics
                'total_assets': total_assets_count,
                'attended_assets': attended_assets_count,
                'remaining_assets': remaining_assets_count,
                'completed_assets': completed_assets,
                'wip_assets': wip_assets,
                'rejected_assets': rejected_assets,
                'rescheduled_assets': rescheduled_assets,
                'not_connected_assets': not_connected_assets,
                'pending_assets': pending_assets,
                
                # Customer metrics (unique assets only)
                'total_customers': total_customers_count,
                'attended_customers': attended_customers_count,
                'remaining_customers': remaining_customers_count,
                
                # Follow-up metrics (only active campaigns)
                'total_followups': total_followups,
                'completed_followups': completed_followups,
                'wip_followups': wip_followups,
                'rejected_followups': rejected_followups,
                'rescheduled_followups': rescheduled_followups,
                'pending_followups': pending_followups,
                'total_quotation_value': total_quotation,
                'quotations_sent_count': quotations_sent,
                'completion_rate': round((completed_followups / total_followups * 100) if total_followups > 0 else 0, 2),
                'time_period': time_period,
                'start_date': start_date,
                'end_date': end_date
            }
            
        except Exception as e:
            print(f"Error in get_performance_summary: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                'total_assets': 0,
                'attended_assets': 0,
                'remaining_assets': 0,
                'completed_assets': 0,
                'wip_assets': 0,
                'rejected_assets': 0,
                'rescheduled_assets': 0,
                'not_connected_assets': 0,
                'pending_assets': 0,
                'total_customers': 0,
                'attended_customers': 0,
                'remaining_customers': 0,
                'total_followups': 0,
                'completed_followups': 0,
                'wip_followups': 0,
                'rejected_followups': 0,
                'rescheduled_followups': 0,
                'pending_followups': 0,
                'total_quotation_value': 0,
                'quotations_sent_count': 0,
                'completion_rate': 0
            }
    
    @staticmethod
    async def export_branch_to_excel(db: Session, branch_code: str, time_period: str = 'all', start_date=None, end_date=None):
        """
        Export branch employee data to Excel with time filter
        """
        try:
            employees = await EmployeePerformanceController.get_employees_by_branch(db, branch_code, time_period, start_date, end_date)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Branch {branch_code} Report"
            
            headers = [
                'User ID', 'Employee Name', 'Total Follow-ups', 'Completed',
                'Work in Progress', 'Rejected', 'Rescheduled', 'Pending', 'Completion Rate %',
                'Quotation Value (₹)', 'Quotations Sent', 'Can Export'
            ]
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            for row_num, emp in enumerate(employees, 2):
                total = emp['total_followups'] or 0
                completed = emp['completed_count'] or 0
                completion_rate = round((completed / total * 100) if total > 0 else 0, 2)
                
                ws.cell(row=row_num, column=1, value=emp['user_id'])
                ws.cell(row=row_num, column=2, value=emp['user_name'])
                ws.cell(row=row_num, column=3, value=total)
                ws.cell(row=row_num, column=4, value=completed)
                ws.cell(row=row_num, column=5, value=emp['wip_count'] or 0)
                ws.cell(row=row_num, column=6, value=emp['rejected_count'] or 0)
                ws.cell(row=row_num, column=7, value=emp.get('rescheduled_count', 0) or 0)
                ws.cell(row=row_num, column=8, value=emp['pending_count'] or 0)
                ws.cell(row=row_num, column=9, value=completion_rate)
                ws.cell(row=row_num, column=10, value=float(emp['total_quotation_value'] or 0))
                ws.cell(row=row_num, column=11, value=emp['quotations_sent_count'] or 0)
                ws.cell(row=row_num, column=12, value="Yes" if emp.get('can_export') else "No")
            
            if start_date and end_date:
                period_text = f"Custom Range: {start_date} to {end_date}"
            else:
                period_text = {
                    'all': 'All Time',
                    'month': 'Last 30 Days',
                    '3months': 'Last 3 Months',
                    '6months': 'Last 6 Months',
                    'year': 'Last 12 Months'
                }.get(time_period, 'All Time')
            
            ws.cell(row=len(employees) + 3, column=1, value=f"Time Period: {period_text}")
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return excel_file
            
        except Exception as e:
            print(f"Error in export_branch_to_excel: {str(e)}")
            import traceback
            traceback.print_exc()
            raise e

    @staticmethod
    async def get_campaign_performance(db: Session, user_info: dict = None, time_period: str = 'all', start_date=None, end_date=None):
        """
        Get campaign performance with time filter applied to follow-ups
        - Attended: ALL customers who have at least one follow-up in this campaign (including those not in asset_numbers)
        - Completed: Total number of completed follow-ups in followup table for this campaign
        - Status and Flag breakdown based on latest follow-up per customer
        """
        try:
            from app.models.campaign_model import Campaign
            from app.models.engagement_model import FollowUp
            from sqlalchemy import and_
            
            campaign_query = db.query(Campaign)
            campaigns = campaign_query.order_by(
                case(
                    (Campaign.status == 'active', 0),
                    else_=1
                ),
                Campaign.created_at.desc()
            ).all()
            
            campaign_performances = []

            # ================= BATCHED PREFETCH (fixes N+1) =================
            # The loop below used to run 6 queries PER campaign. All of them are
            # keyed by campaign_id, so run them ONCE for all campaigns (chunked
            # IN() to stay well under MSSQL's 2100-parameter cap) and build
            # per-campaign maps consumed inside the loop.
            all_campaign_ids = [c.id for c in campaigns]
            followup_only_by_campaign = {}    # cid -> set(instance_id) status != completed (NO time filter, as before)
            attended_by_campaign = {}         # cid -> set(instance_id) with time filter
            followups_by_campaign = {}        # cid -> [FollowUp rows] (instance not null, time filter)
            completed_count_by_campaign = {}  # cid -> count(status == 'completed', time filter)

            CHUNK = 1000
            for i in range(0, len(all_campaign_ids), CHUNK):
                chunk = all_campaign_ids[i:i + CHUNK]

                # Same filters as the old per-campaign "followup_only_customers_query"
                followup_only_rows = db.query(
                    FollowUp.campaign_id, FollowUp.customer_instance_id
                ).filter(
                    FollowUp.campaign_id.in_(chunk),
                    FollowUp.customer_instance_id.isnot(None),
                    FollowUp.status != 'completed'  # Ignore completed customers
                ).distinct().all()
                for cid, iid in followup_only_rows:
                    if iid:
                        followup_only_by_campaign.setdefault(cid, set()).add(iid)

                # Same filters as the old per-campaign "all_customers_with_followups_query"
                attended_query = db.query(
                    FollowUp.campaign_id, FollowUp.customer_instance_id
                ).filter(
                    FollowUp.campaign_id.in_(chunk),
                    FollowUp.customer_instance_id.isnot(None)
                )
                attended_query = EmployeePerformanceController.apply_time_filter(
                    attended_query, FollowUp, time_period, start_date, end_date
                ).distinct()
                for cid, iid in attended_query.all():
                    if iid:
                        attended_by_campaign.setdefault(cid, set()).add(iid)

                # Same filters as the old per-campaign "filtered_query" (STEP 2)
                followups_query = db.query(FollowUp).filter(
                    FollowUp.campaign_id.in_(chunk),
                    FollowUp.customer_instance_id.isnot(None)
                )
                followups_query = EmployeePerformanceController.apply_time_filter(
                    followups_query, FollowUp, time_period, start_date, end_date
                )
                for fu in followups_query.all():
                    followups_by_campaign.setdefault(fu.campaign_id, []).append(fu)

                # Same filters as the old per-campaign completed COUNT (STEPs 4 & 6)
                completed_query = db.query(
                    FollowUp.campaign_id, func.count(FollowUp.id)
                ).filter(
                    FollowUp.campaign_id.in_(chunk),
                    FollowUp.status == 'completed'
                )
                completed_query = EmployeePerformanceController.apply_time_filter(
                    completed_query, FollowUp, time_period, start_date, end_date
                ).group_by(FollowUp.campaign_id)
                for cid, cnt in completed_query.all():
                    completed_count_by_campaign[cid] = int(cnt or 0)
            # ================ END BATCHED PREFETCH ================

            for campaign in campaigns:
                asset_numbers = campaign.asset_numbers or []
                asset_numbers_set = set(asset_numbers)

                # Get unique instance_ids from followups table for this campaign where status is NOT completed
                followup_only_customers = followup_only_by_campaign.get(campaign.id, set())

                # Remove those that are already in asset_numbers
                asset_numbers_set = set(asset_numbers)
                additional_customers = followup_only_customers - asset_numbers_set

                # Total remaining = asset_numbers_count + additional customers from followups (excluding completed)
                remaining_count = len(asset_numbers) + len(additional_customers)

                # ========== STEP 1: Get customers with follow-ups in this campaign - WITH TIME FILTER ==========
                attended_customers_count = len(attended_by_campaign.get(campaign.id, set()))

                # ========== STEP 2: Get latest follow-up for each customer (with time filter) ==========
                # All follow-ups for this campaign within time period (prefetched above)
                all_followups = followups_by_campaign.get(campaign.id, [])
                
                # Track latest follow-up per customer (by customer_instance_id)
                latest_followup_map = {}
                for fu in all_followups:
                    instance_id = fu.customer_instance_id
                    if instance_id:
                        followup_date = fu.followup_date if fu.followup_date else fu.created_at
                        if instance_id not in latest_followup_map:
                            latest_followup_map[instance_id] = fu
                        else:
                            existing_date = latest_followup_map[instance_id].followup_date if latest_followup_map[instance_id].followup_date else latest_followup_map[instance_id].created_at
                            if followup_date > existing_date:
                                latest_followup_map[instance_id] = fu
                
                # ========== STEP 3: Count status breakdown based on latest follow-up ==========
                wip_count = 0
                rejected_count = 0
                rescheduled_count = 0
                pending_count = 0
                not_connected_count = 0  # Customers whose latest status is not_connected
                completed_in_latest = 0  # Customers whose latest status is completed
                
                # Flag breakdown based on latest follow-up
                flag_counts = {}
                
                for instance_id, latest_fu in latest_followup_map.items():
                    status = latest_fu.status or 'pending'
                    if status == 'completed':
                        completed_in_latest += 1
                    elif status == 'wip':
                        wip_count += 1
                    elif status == 'rejected':
                        rejected_count += 1
                    elif status == 'rescheduled':
                        rescheduled_count += 1
                    elif status == 'not_connected':
                        not_connected_count += 1
                    elif status == 'pending':
                        pending_count += 1
                    
                    # Count flags
                    flag = latest_fu.followup_flag
                    if flag:
                        flag_counts[flag] = flag_counts.get(flag, 0) + 1
                
                # ========== STEP 4: Total completed follow-ups (all completed records, not just latest) ==========
                # Prefetched grouped COUNT (same filters: campaign + completed + time filter)
                total_completed_followups = completed_count_by_campaign.get(campaign.id, 0)

                # ========== STEP 5: Total follow-ups in time period ==========
                # Same filters as the prefetched follow-up list -> its length IS the count
                total_followups = len(all_followups)

                # ========== STEP 6: Total customers (asset_numbers + period completed) ==========
                # Identical query to STEP 4 -> reuse the same prefetched count
                period_completed = completed_count_by_campaign.get(campaign.id, 0)

                total_customers = remaining_count + period_completed
                
                # ========== STEP 7: Success percentage ==========
                success_percentage = round((total_completed_followups / total_followups * 100) if total_followups > 0 else 0, 2)
                
                campaign_performances.append({
                    'campaign_id': campaign.id,
                    'campaign_name': campaign.name,
                    'service': campaign.service,
                    'description': campaign.description,
                    'color': campaign.color or '#71C9CE',
                    'status': campaign.status,
                    'start_date': campaign.start_date.isoformat() if campaign.start_date else None,
                    'end_date': campaign.end_date.isoformat() if campaign.end_date else None,
                    'created_by_name': campaign.created_by_name,
                    'asset_numbers_count': remaining_count,  # Remaining customers (from asset_numbers)
                    'attended_customers': attended_customers_count,  # ALL customers with at least one follow-up
                    'total_completed_followups': total_completed_followups,  # Total completed in followup table
                    'completed_in_latest': completed_in_latest,  # Customers with latest status = completed
                    'wip_count': wip_count,  # Customers with latest status = wip
                    'rejected_count': rejected_count,  # Customers with latest status = rejected
                    'rescheduled_count': rescheduled_count,  # Customers with latest status = rescheduled
                    'not_connected_count': not_connected_count,  # Customers with latest status = not_connected
                    'pending_count': pending_count,  # Customers with latest status = pending
                    'total_followups': total_followups,  # Total follow-up records in time period
                    'total_customers': total_customers,
                    'success_percentage': success_percentage,
                    'flag_breakdown': flag_counts,  # Based on latest follow-up
                    'time_period': time_period,
                    'start_date': start_date,
                    'end_date': end_date
                })
            
            return campaign_performances
            
        except Exception as e:
            print(f"Error in get_campaign_performance: {str(e)}")
            import traceback
            traceback.print_exc()
            return []
    
    @staticmethod
    async def get_campaign_details(db: Session, campaign_id: int, user_info: dict = None, time_period: str = 'all', start_date=None, end_date=None):
        """
        Get detailed performance for a specific campaign with time filter
        """
        try:
            from app.models.campaign_model import Campaign
            from app.models.engagement_model import FollowUp
            
            campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
            if not campaign:
                return None
            
            base_query = db.query(FollowUp).filter(FollowUp.campaign_id == campaign_id)
            filtered_query = EmployeePerformanceController.apply_time_filter(base_query, FollowUp, time_period, start_date, end_date)
            
            followup_stats = filtered_query.with_entities(
                func.count(FollowUp.id).label('total'),
                func.sum(case((FollowUp.status == 'completed', 1), else_=0)).label('completed'),
                func.sum(case((FollowUp.status == 'wip', 1), else_=0)).label('wip'),
                func.sum(case((FollowUp.status == 'rejected', 1), else_=0)).label('rejected'),
                func.sum(case((FollowUp.status == 'pending', 1), else_=0)).label('pending'),
                func.sum(case((FollowUp.status == 'rescheduled', 1), else_=0)).label('rescheduled'),
                # Only count quotation when status is 'completed'
                func.sum(case((and_(FollowUp.quotation_sent == True, FollowUp.status == 'completed'), FollowUp.quotation_value), else_=0)).label('total_quotation'),
                func.sum(case((and_(FollowUp.quotation_sent == True, FollowUp.status == 'completed'), 1), else_=0)).label('quotations_sent')
            ).first()
            
            user_breakdown_raw = filtered_query.with_entities(
                FollowUp.user_id,
                FollowUp.user_name,
                func.count(FollowUp.id).label('total'),
                func.sum(case((FollowUp.status == 'completed', 1), else_=0)).label('completed')
            ).group_by(
                FollowUp.user_id, FollowUp.user_name
            ).all()
            
            user_breakdown = []
            for item in user_breakdown_raw:
                total = int(item.total) if item.total else 0
                completed = int(item.completed) if item.completed else 0
                user_breakdown.append({
                    'user_id': item.user_id,
                    'user_name': item.user_name,
                    'total_followups': total,
                    'completed_count': completed,
                    'success_rate': round((completed / total * 100) if total > 0 else 0, 2)
                })
            
            user_breakdown.sort(key=lambda x: x['completed_count'], reverse=True)
            
            total = int(followup_stats.total) if followup_stats.total else 0
            completed = int(followup_stats.completed) if followup_stats.completed else 0
            
            return {
                'campaign_id': campaign.id,
                'campaign_name': campaign.name,
                'service': campaign.service,
                'description': campaign.description,
                'color': campaign.color or '#71C9CE',
                'status': campaign.status,
                'start_date': campaign.start_date.isoformat() if campaign.start_date else None,
                'end_date': campaign.end_date.isoformat() if campaign.end_date else None,
                'created_by_name': campaign.created_by_name,
                'asset_numbers_count': len(campaign.asset_numbers or []),
                'total_followups': total,
                'completed_count': completed,
                'wip_count': int(followup_stats.wip) if followup_stats.wip else 0,
                'rejected_count': int(followup_stats.rejected) if followup_stats.rejected else 0,
                'rescheduled_count': int(followup_stats.rescheduled) if followup_stats.rescheduled else 0,
                'pending_count': int(followup_stats.pending) if followup_stats.pending else 0,
                'total_quotation_value': float(followup_stats.total_quotation) if followup_stats.total_quotation else 0,
                'quotations_sent_count': int(followup_stats.quotations_sent) if followup_stats.quotations_sent else 0,
                'success_percentage': round((completed / total * 100) if total > 0 else 0, 2),
                'user_breakdown': user_breakdown,
                'time_period': time_period,
                'start_date': start_date,
                'end_date': end_date
            }
            
        except Exception as e:
            print(f"Error in get_campaign_details: {str(e)}")
            import traceback
            traceback.print_exc()
            return None     

    @staticmethod
    async def get_activity_stats(
        db: Session, 
        campaign_ids: List[int] = None, 
        time_period: str = 'all', 
        start_date=None, 
        end_date=None,
        user_role: str = None,  # ADD THIS
        user_branch: str = None  # ADD THIS
    ):
        """
        Get activity statistics with counts and time filter, optionally filtered by campaigns
        Includes additional customers from followups (excluding completed status)
        """
        try:
            from app.models.engagement_model import Activity, FollowUp
            from app.models.user_model import User
            from app.models.campaign_model import Campaign
            
            all_activities = db.query(Activity).all()
            
            campaign_totals = {
                'total_customers': 0,
                'total_followups': 0
            }
            
            if campaign_ids and len(campaign_ids) > 0:
                campaigns = db.query(Campaign).filter(Campaign.id.in_(campaign_ids)).all()
            else:
                campaigns = db.query(Campaign).filter(Campaign.status == 'active').all()

            # ===== BATCHED PREFETCH (fixes N+1: was 3 queries per campaign) =====
            selected_campaign_ids = [c.id for c in campaigns]
            followup_only_by_campaign = {}   # cid -> set(instance_id), status != completed, time filter
            completed_count_by_campaign = {} # cid -> completed count, time filter
            total_count_by_campaign = {}     # cid -> total follow-up count, time filter

            CHUNK = 1000
            for i in range(0, len(selected_campaign_ids), CHUNK):
                chunk = selected_campaign_ids[i:i + CHUNK]

                followup_only_query = db.query(
                    FollowUp.campaign_id, FollowUp.customer_instance_id
                ).filter(
                    FollowUp.campaign_id.in_(chunk),
                    FollowUp.customer_instance_id.isnot(None),
                    FollowUp.status != 'completed'
                ).distinct()
                followup_only_query = EmployeePerformanceController.apply_time_filter(
                    followup_only_query, FollowUp, time_period, start_date, end_date
                )
                for cid, iid in followup_only_query.all():
                    if iid:
                        followup_only_by_campaign.setdefault(cid, set()).add(iid)

                completed_count_query = db.query(
                    FollowUp.campaign_id, func.count(FollowUp.id)
                ).filter(
                    FollowUp.campaign_id.in_(chunk),
                    FollowUp.status == 'completed'
                )
                completed_count_query = EmployeePerformanceController.apply_time_filter(
                    completed_count_query, FollowUp, time_period, start_date, end_date
                ).group_by(FollowUp.campaign_id)
                for cid, cnt in completed_count_query.all():
                    completed_count_by_campaign[cid] = int(cnt or 0)

                total_count_query = db.query(
                    FollowUp.campaign_id, func.count(FollowUp.id)
                ).filter(
                    FollowUp.campaign_id.in_(chunk)
                )
                total_count_query = EmployeePerformanceController.apply_time_filter(
                    total_count_query, FollowUp, time_period, start_date, end_date
                ).group_by(FollowUp.campaign_id)
                for cid, cnt in total_count_query.all():
                    total_count_by_campaign[cid] = int(cnt or 0)
            # ===== END BATCHED PREFETCH =====

            for campaign in campaigns:
                asset_numbers = campaign.asset_numbers or []
                asset_numbers_set = set(asset_numbers)
                asset_numbers_count = len(asset_numbers)

                # Get additional customers from followups - WITH TIME FILTER (prefetched)
                followup_only_customers = followup_only_by_campaign.get(campaign.id, set())
                additional_customers = followup_only_customers - asset_numbers_set

                # completed_count with time filter (prefetched)
                completed_count = completed_count_by_campaign.get(campaign.id, 0)

                campaign_totals['total_customers'] += asset_numbers_count + completed_count + len(additional_customers)

                campaign_totals['total_followups'] += total_count_by_campaign.get(campaign.id, 0)

            # Start building the main query
            base_query = db.query(FollowUp).filter(FollowUp.activity_id.isnot(None))

            if campaign_ids and len(campaign_ids) > 0:
                base_query = base_query.filter(FollowUp.campaign_id.in_(campaign_ids))
            else:
                # `campaigns` already holds the active campaigns here (fetched above
                # with the exact same filter) — no need to re-query.
                active_campaign_ids = [c.id for c in campaigns]
                if active_campaign_ids:
                    base_query = base_query.filter(FollowUp.campaign_id.in_(active_campaign_ids))
            
            # APPLY BRANCH FILTER FOR BRANCH ADMIN
            if user_role and user_role.lower() == 'branch_admin' and user_branch:
                # Get all users in this branch
                branch_users = db.query(User.user_id).filter(User.branch == user_branch).all()
                branch_user_ids = [u[0] for u in branch_users]
                
                if branch_user_ids:
                    base_query = base_query.filter(FollowUp.user_id.in_(branch_user_ids))
                else:
                    # No users in this branch, return empty result
                    return {'activities': [], 'campaign_totals': campaign_totals}
            
            filtered_query = EmployeePerformanceController.apply_time_filter(base_query, FollowUp, time_period, start_date, end_date)
            
            activity_counts_raw = filtered_query.with_entities(
                FollowUp.activity_id,
                FollowUp.user_id,
                FollowUp.user_name,
                func.count(FollowUp.id).label('count')
            ).group_by(FollowUp.activity_id, FollowUp.user_id, FollowUp.user_name).all()
            
            # Only fetch users referenced in activity entries — and only the columns we need
            referenced_user_ids = {item.user_id for item in activity_counts_raw if item.user_id}
            user_branch_map = {}
            user_branch_name_map = {}
            if referenced_user_ids:
                users_rows = db.query(User.user_id, User.branch, User.branch_name).filter(
                    User.user_id.in_(referenced_user_ids)
                ).all()
                user_branch_map = {u.user_id: u.branch for u in users_rows}
                user_branch_name_map = {u.user_id: u.branch_name for u in users_rows}
            
            # Pre-group once (was a full scan of activity_counts_raw per activity);
            # per-activity entry order is unchanged (encounter order).
            entries_by_activity = {}
            for item in activity_counts_raw:
                entries_by_activity.setdefault(item.activity_id, []).append(item)

            activity_stats = []
            for activity in all_activities:
                activity_entries = entries_by_activity.get(activity.id, [])
                
                user_breakdown = []
                total_count = 0
                
                for entry in activity_entries:
                    branch = user_branch_map.get(entry.user_id, 'Unknown')
                    branch_name = user_branch_name_map.get(entry.user_id, branch)
                    count = int(entry.count)
                    total_count += count
                    
                    user_breakdown.append({
                        'user_id': entry.user_id,
                        'user_name': entry.user_name,
                        'branch': branch,
                        'branch_display': branch_name or branch,
                        'count': count
                    })
                
                user_breakdown.sort(key=lambda x: x['count'], reverse=True)
                
                if total_count > 0:
                    activity_stats.append({
                        'activity_id': activity.id,
                        'activity_name': activity.content,
                        'total_count': total_count,
                        'user_breakdown': user_breakdown
                    })
            
            activity_stats.sort(key=lambda x: x['total_count'], reverse=True)
            
            return {
                'activities': activity_stats,
                'campaign_totals': campaign_totals
            }
            
        except Exception as e:
            print(f"Error in get_activity_stats: {str(e)}")
            import traceback
            traceback.print_exc()
            return {'activities': [], 'campaign_totals': {'total_customers': 0, 'total_followups': 0}}

    @staticmethod
    async def get_rr_stats(
        db: Session, 
        campaign_ids: List[int] = None, 
        time_period: str = 'all', 
        start_date=None, 
        end_date=None,
        user_role: str = None,  # ADD THIS
        user_branch: str = None  # ADD THIS
    ):
        """
        Get RR (Rejected Reason) statistics with counts and time filter, optionally filtered by campaigns
        Includes additional customers from followups (excluding completed status)
        """
        try:
            from app.models.engagement_model import RR, FollowUp
            from app.models.user_model import User
            from app.models.campaign_model import Campaign
            
            all_rrs = db.query(RR).all()
            
            campaign_totals = {
                'total_customers': 0,
                'total_followups': 0
            }
            
            if campaign_ids and len(campaign_ids) > 0:
                campaigns = db.query(Campaign).filter(Campaign.id.in_(campaign_ids)).all()
            else:
                campaigns = db.query(Campaign).filter(Campaign.status == 'active').all()
            
            # ===== BATCHED PREFETCH (fixes N+1: was 3 queries per campaign) =====
            selected_campaign_ids = [c.id for c in campaigns]
            followup_only_by_campaign = {}   # cid -> set(instance_id), status != completed, time filter
            completed_count_by_campaign = {} # cid -> completed count, time filter
            total_count_by_campaign = {}     # cid -> total follow-up count, time filter

            CHUNK = 1000
            for i in range(0, len(selected_campaign_ids), CHUNK):
                chunk = selected_campaign_ids[i:i + CHUNK]

                followup_only_query = db.query(
                    FollowUp.campaign_id, FollowUp.customer_instance_id
                ).filter(
                    FollowUp.campaign_id.in_(chunk),
                    FollowUp.customer_instance_id.isnot(None),
                    FollowUp.status != 'completed'
                ).distinct()
                followup_only_query = EmployeePerformanceController.apply_time_filter(
                    followup_only_query, FollowUp, time_period, start_date, end_date
                )
                for cid, iid in followup_only_query.all():
                    if iid:
                        followup_only_by_campaign.setdefault(cid, set()).add(iid)

                completed_count_query = db.query(
                    FollowUp.campaign_id, func.count(FollowUp.id)
                ).filter(
                    FollowUp.campaign_id.in_(chunk),
                    FollowUp.status == 'completed'
                )
                completed_count_query = EmployeePerformanceController.apply_time_filter(
                    completed_count_query, FollowUp, time_period, start_date, end_date
                ).group_by(FollowUp.campaign_id)
                for cid, cnt in completed_count_query.all():
                    completed_count_by_campaign[cid] = int(cnt or 0)

                total_count_query = db.query(
                    FollowUp.campaign_id, func.count(FollowUp.id)
                ).filter(
                    FollowUp.campaign_id.in_(chunk)
                )
                total_count_query = EmployeePerformanceController.apply_time_filter(
                    total_count_query, FollowUp, time_period, start_date, end_date
                ).group_by(FollowUp.campaign_id)
                for cid, cnt in total_count_query.all():
                    total_count_by_campaign[cid] = int(cnt or 0)
            # ===== END BATCHED PREFETCH =====

            for campaign in campaigns:
                asset_numbers = campaign.asset_numbers or []
                asset_numbers_set = set(asset_numbers)
                asset_numbers_count = len(asset_numbers)

                # Get additional customers from followups - WITH TIME FILTER (prefetched)
                followup_only_customers = followup_only_by_campaign.get(campaign.id, set())
                additional_customers = followup_only_customers - asset_numbers_set

                # completed_count with time filter (prefetched)
                completed_count = completed_count_by_campaign.get(campaign.id, 0)

                campaign_totals['total_customers'] += asset_numbers_count + completed_count + len(additional_customers)

                campaign_totals['total_followups'] += total_count_by_campaign.get(campaign.id, 0)

            # Start building the main query
            base_query = db.query(FollowUp).filter(FollowUp.rr_id.isnot(None))

            if campaign_ids and len(campaign_ids) > 0:
                base_query = base_query.filter(FollowUp.campaign_id.in_(campaign_ids))
            else:
                # `campaigns` already holds the active campaigns here (fetched above
                # with the exact same filter) — no need to re-query.
                active_campaign_ids = [c.id for c in campaigns]
                if active_campaign_ids:
                    base_query = base_query.filter(FollowUp.campaign_id.in_(active_campaign_ids))
            
            # APPLY BRANCH FILTER FOR BRANCH ADMIN
            if user_role and user_role.lower() == 'branch_admin' and user_branch:
                # Get all users in this branch
                branch_users = db.query(User.user_id).filter(User.branch == user_branch).all()
                branch_user_ids = [u[0] for u in branch_users]
                
                if branch_user_ids:
                    base_query = base_query.filter(FollowUp.user_id.in_(branch_user_ids))
                else:
                    # No users in this branch, return empty result
                    return {'rr_reasons': [], 'campaign_totals': campaign_totals}
            
            filtered_query = EmployeePerformanceController.apply_time_filter(base_query, FollowUp, time_period, start_date, end_date)
            
            rr_counts_raw = filtered_query.with_entities(
                FollowUp.rr_id,
                FollowUp.user_id,
                FollowUp.user_name,
                func.count(FollowUp.id).label('count')
            ).group_by(FollowUp.rr_id, FollowUp.user_id, FollowUp.user_name).all()
            
            # Only fetch users referenced in rr entries — and only the columns we need
            referenced_user_ids = {item.user_id for item in rr_counts_raw if item.user_id}
            user_branch_map = {}
            user_branch_name_map = {}
            if referenced_user_ids:
                users_rows = db.query(User.user_id, User.branch, User.branch_name).filter(
                    User.user_id.in_(referenced_user_ids)
                ).all()
                user_branch_map = {u.user_id: u.branch for u in users_rows}
                user_branch_name_map = {u.user_id: u.branch_name for u in users_rows}
            
            # Pre-group once (was a full scan of rr_counts_raw per RR);
            # per-RR entry order is unchanged (encounter order).
            entries_by_rr = {}
            for item in rr_counts_raw:
                entries_by_rr.setdefault(item.rr_id, []).append(item)

            rr_stats = []
            for rr in all_rrs:
                rr_entries = entries_by_rr.get(rr.id, [])
                
                user_breakdown = []
                total_count = 0
                
                for entry in rr_entries:
                    branch = user_branch_map.get(entry.user_id, 'Unknown')
                    branch_name = user_branch_name_map.get(entry.user_id, branch)
                    count = int(entry.count)
                    total_count += count
                    
                    user_breakdown.append({
                        'user_id': entry.user_id,
                        'user_name': entry.user_name,
                        'branch': branch,
                        'branch_display': branch_name or branch,
                        'count': count
                    })
                
                user_breakdown.sort(key=lambda x: x['count'], reverse=True)
                
                if total_count > 0:
                    rr_stats.append({
                        'rr_id': rr.id,
                        'rr_name': rr.content,
                        'total_count': total_count,
                        'user_breakdown': user_breakdown
                    })
            
            rr_stats.sort(key=lambda x: x['total_count'], reverse=True)
            
            return {
                'rr_reasons': rr_stats,
                'campaign_totals': campaign_totals
            }
            
        except Exception as e:
            print(f"Error in get_rr_stats: {str(e)}")
            import traceback
            traceback.print_exc()
            return {'rr_reasons': [], 'campaign_totals': {'total_customers': 0, 'total_followups': 0}}

    @staticmethod
    async def get_all_campaigns_list(db: Session, user_role: str = None, user_branch: str = None):
        """
        Get all ACTIVE campaigns for filter dropdown
        For branch_admin, only show campaigns that have customers from their branch
        """
        try:
            from app.models.campaign_model import Campaign
            from app.models.customer_model import Customer
            
            # Get all active campaigns
            campaigns = db.query(Campaign).filter(
                Campaign.status == 'active'
            ).order_by(Campaign.name).all()
            
            #commented by nik
            # If branch_admin, filter campaigns that have customers in their branch
            # if user_role and user_role.lower() == 'branch_admin' and user_branch:
            #     filtered_campaigns = []
                
            #     for campaign in campaigns:
            #         asset_numbers = campaign.asset_numbers or []
            #         if asset_numbers:
            #             # Check if any customer from this branch is in asset_numbers
            #             branch_customers = db.query(Customer).filter(
            #                 Customer.instance_id.in_(asset_numbers),
            #                 Customer.branch_id == user_branch
            #             ).first()
                        
            #             if branch_customers:
            #                 filtered_campaigns.append(campaign)
            #                 continue
                    
            #         # Also check if there are any follow-ups from this branch for this campaign
            #         from app.models.engagement_model import FollowUp
            #         from app.models.user_model import User
                    
            #         branch_users = db.query(User.user_id).filter(User.branch == user_branch).all()
            #         branch_user_ids = [u[0] for u in branch_users]
                    
            #         if branch_user_ids:
            #             branch_followups = db.query(FollowUp).filter(
            #                 FollowUp.campaign_id == campaign.id,
            #                 FollowUp.user_id.in_(branch_user_ids)
            #             ).first()
                        
            #             if branch_followups:
            #                 filtered_campaigns.append(campaign)
                
            #     campaigns = filtered_campaigns

            #added by nik
            # If branch_admin, filter campaigns that have customers in their branch
            if user_role and user_role.lower() == 'branch_admin' and user_branch:
                from app.models.engagement_model import FollowUp
                from app.models.user_model import User

                # Pull this branch's customer instance_ids ONCE (indexed query) and
                # do the asset_numbers membership test in Python. The old code ran
                # Customer.instance_id.in_(asset_numbers) per campaign, and
                # asset_numbers can hold 2600+ ids — past SQL Server's 2100-param
                # IN() limit, which is what raised the 07002 error.
                branch_rows = db.query(Customer.instance_id).filter(
                    Customer.branch_id == user_branch
                ).all()
                branch_instance_ids = {r[0] for r in branch_rows if r[0]}

                # Branch user_ids fetched ONCE (was re-queried inside the loop).
                branch_user_ids = [
                    u[0] for u in db.query(User.user_id)
                    .filter(User.branch == user_branch).all()
                ]

                # Campaign ids that already have a follow-up from any branch user.
                campaigns_with_branch_followups = set()
                if branch_user_ids:
                    rows = db.query(FollowUp.campaign_id).filter(
                        FollowUp.user_id.in_(branch_user_ids),
                        FollowUp.campaign_id.isnot(None)
                    ).distinct().all()
                    campaigns_with_branch_followups = {r[0] for r in rows}

                filtered_campaigns = []
                for campaign in campaigns:
                    asset_set = set(campaign.asset_numbers or [])
                    if asset_set and (branch_instance_ids & asset_set):
                        filtered_campaigns.append(campaign)
                    elif campaign.id in campaigns_with_branch_followups:
                        filtered_campaigns.append(campaign)

                campaigns = filtered_campaigns
            
            result = [
                {
                    'id': campaign.id,
                    'name': campaign.name,
                    'service': campaign.service,
                    'status': campaign.status
                }
                for campaign in campaigns
            ]
            
            return result
            
        except Exception as e:
            print(f"Error in get_all_campaigns_list: {str(e)}")
            import traceback
            traceback.print_exc()
            return []
    
    @staticmethod
    async def get_branch_top_campaigns(db: Session, branch_code: str, time_period: str = 'all', start_date=None, end_date=None):
        """
        Get top 3 campaigns by completed count for a specific branch with time filter
        """
        try:
            from app.models.campaign_model import Campaign
            from app.models.engagement_model import FollowUp
            from app.models.user_model import User
            
            users = db.query(User).filter(
                User.branch == branch_code,
                User.is_blocked == False
            ).all()
            
            user_ids = [u.user_id for u in users]
            
            if not user_ids:
                return []
            
            base_query = db.query(
                FollowUp.campaign_id,
                func.count(FollowUp.id).label('completed_count')
            ).filter(
                FollowUp.user_id.in_(user_ids),
                FollowUp.status == 'completed',
                FollowUp.campaign_id.isnot(None)
            )
            
            filtered_query = EmployeePerformanceController.apply_time_filter(base_query, FollowUp, time_period, start_date, end_date)
            
            campaign_counts = filtered_query.group_by(FollowUp.campaign_id).all()
            
            # Batch-fetch all referenced campaigns in ONE query instead of N+1
            referenced_ids = [item.campaign_id for item in campaign_counts if item.campaign_id]
            campaign_map = {}
            if referenced_ids:
                for c in db.query(Campaign).filter(Campaign.id.in_(referenced_ids)).all():
                    campaign_map[c.id] = c
            
            branch_campaigns = []
            for item in campaign_counts:
                if item.campaign_id:
                    campaign = campaign_map.get(item.campaign_id)
                    if campaign:
                        branch_campaigns.append({
                            'campaign_id': campaign.id,
                            'campaign_name': campaign.name,
                            'service': campaign.service,
                            'completed_count': int(item.completed_count)
                        })
            
            branch_campaigns.sort(key=lambda x: x['completed_count'], reverse=True)
            
            return branch_campaigns[:3]
            
        except Exception as e:
            print(f"Error in get_branch_top_campaigns: {str(e)}")
            import traceback
            traceback.print_exc()
            return []
        
    @staticmethod
    async def get_campaign_customers(db: Session, campaign_id: int, user_info: dict = None):
        """
        Get all customers for a specific campaign with their latest follow-up details
        Includes customers from asset_numbers AND customers from followups table (excluding completed status)
        """
        try:
            from app.models.campaign_model import Campaign
            from app.models.engagement_model import FollowUp
            from app.models.customer_model import Customer
            
            campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
            if not campaign:
                return None
            
            asset_numbers = campaign.asset_numbers or []
            asset_numbers_set = set(asset_numbers)
            
            # ========== STEP 1: Get customers from asset_numbers ==========
            customers_from_assets = db.query(Customer).filter(Customer.instance_id.in_(asset_numbers)).all()
            customer_map = {c.instance_id: c for c in customers_from_assets}
            
            # ========== STEP 2: Get additional customers from followups table (not in asset_numbers and status NOT completed) ==========
            # Get unique instance_ids from followups for this campaign where status is NOT completed
            followup_customers_query = db.query(FollowUp.customer_instance_id).filter(
                FollowUp.campaign_id == campaign_id,
                FollowUp.customer_instance_id.isnot(None),
                FollowUp.status != 'completed'  # EXCLUDE completed status customers
            ).distinct()
            
            followup_instance_ids = set([row[0] for row in followup_customers_query.all() if row[0]])
            
            # Remove those already in asset_numbers
            additional_instance_ids = followup_instance_ids - asset_numbers_set
            
            # Get customer details for additional instance_ids
            additional_customers = []
            if additional_instance_ids:
                additional_customers = db.query(Customer).filter(Customer.instance_id.in_(list(additional_instance_ids))).all()
                for cust in additional_customers:
                    customer_map[cust.instance_id] = cust
            
            # For instance_ids not found in customers table, create placeholder
            all_instance_ids = list(asset_numbers_set) + list(additional_instance_ids)
            
            # ========== STEP 3: Get latest follow-up for each customer in this campaign ==========
            from sqlalchemy import func, and_
            
            # Subquery to get latest follow-up per customer_instance_id
            latest_followup_subquery = db.query(
                FollowUp.customer_instance_id,
                func.max(FollowUp.created_at).label('latest_date')
            ).filter(
                FollowUp.campaign_id == campaign_id,
                FollowUp.customer_instance_id.in_(all_instance_ids) if all_instance_ids else FollowUp.customer_instance_id.isnot(None)
            ).group_by(FollowUp.customer_instance_id).subquery()
            
            latest_followups = db.query(FollowUp).join(
                latest_followup_subquery,
                and_(
                    FollowUp.customer_instance_id == latest_followup_subquery.c.customer_instance_id,
                    FollowUp.created_at == latest_followup_subquery.c.latest_date
                )
            ).all()
            
            followup_map = {}
            for fu in latest_followups:
                followup_map[fu.customer_instance_id] = fu
            
            # ========== STEP 4: Build customers list ==========
            customers_list = []
            
            # First, add customers from asset_numbers
            for instance_id in asset_numbers:
                customer = customer_map.get(instance_id)
                followup = followup_map.get(instance_id)
                
                if customer:
                    customer_data = {
                        's_no': 0,
                        'instance_id': customer.instance_id,
                        'customer_name': customer.customer_name or 'N/A',
                        'phone_number': customer.phone_number or 'N/A',
                        'email': customer.email or 'N/A',
                        'branch_id': customer.branch_id or 'N/A',
                        'location': customer.location or 'N/A',
                        'last_status': followup.status if followup else '-',
                        'last_followup_user_name': followup.user_name if followup else 'N/A',
                        'last_followup_user_id': followup.user_id if followup else 'N/A',
                        'last_followup_date': followup.created_at.isoformat() if followup and followup.created_at else None,
                        'next_followup_date': followup.next_followup_date.isoformat() if followup and followup.next_followup_date else None,
                        'latest_flag': followup.followup_flag if followup else 'N/A',
                        'latest_remark': followup.followup_remark if followup else 'N/A',
                        'quotation_sent': followup.quotation_sent if followup else False,
                        'quotation_value': followup.quotation_value if followup else 0,
                        'quotation_no': followup.quotation_no if followup else None,
                        'source': 'asset_numbers'
                    }
                else:
                    customer_data = {
                        's_no': 0,
                        'instance_id': instance_id,
                        'customer_name': 'Not Found in Database',
                        'phone_number': 'N/A',
                        'email': 'N/A',
                        'branch_id': 'N/A',
                        'location': 'N/A',
                        'last_status': followup.status if followup else '-',
                        'last_followup_user_name': followup.user_name if followup else 'N/A',
                        'last_followup_user_id': followup.user_id if followup else 'N/A',
                        'last_followup_date': followup.created_at.isoformat() if followup and followup.created_at else None,
                        'next_followup_date': followup.next_followup_date.isoformat() if followup and followup.next_followup_date else None,
                        'latest_flag': followup.followup_flag if followup else 'N/A',
                        'latest_remark': followup.followup_remark if followup else 'N/A',
                        'quotation_sent': followup.quotation_sent if followup else False,
                        'quotation_value': followup.quotation_value if followup else 0,
                        'quotation_no': followup.quotation_no if followup else None,
                        'source': 'asset_numbers'
                    }
                
                customers_list.append(customer_data)
            
            # Second, add additional customers from followups (not in asset_numbers and status not completed)
            for instance_id in additional_instance_ids:
                # Skip if already added from asset_numbers
                if instance_id in asset_numbers_set:
                    continue
                    
                customer = customer_map.get(instance_id)
                followup = followup_map.get(instance_id)
                
                if customer:
                    customer_data = {
                        's_no': 0,
                        'instance_id': customer.instance_id,
                        'customer_name': customer.customer_name or 'N/A',
                        'phone_number': customer.phone_number or 'N/A',
                        'email': customer.email or 'N/A',
                        'branch_id': customer.branch_id or 'N/A',
                        'location': customer.location or 'N/A',
                        'last_status': followup.status if followup else '-',
                        'last_followup_user_name': followup.user_name if followup else 'N/A',
                        'last_followup_user_id': followup.user_id if followup else 'N/A',
                        'last_followup_date': followup.created_at.isoformat() if followup and followup.created_at else None,
                        'next_followup_date': followup.next_followup_date.isoformat() if followup and followup.next_followup_date else None,
                        'latest_flag': followup.followup_flag if followup else 'N/A',
                        'latest_remark': followup.followup_remark if followup else 'N/A',
                        'quotation_sent': followup.quotation_sent if followup else False,
                        'quotation_value': followup.quotation_value if followup else 0,
                        'quotation_no': followup.quotation_no if followup else None,
                        'source': 'followup_only'
                    }
                else:
                    customer_data = {
                        's_no': 0,
                        'instance_id': instance_id,
                        'customer_name': 'Customer Not Found',
                        'phone_number': 'N/A',
                        'email': 'N/A',
                        'branch_id': 'N/A',
                        'location': 'N/A',
                        'last_status': followup.status if followup else '-',
                        'last_followup_user_name': followup.user_name if followup else 'N/A',
                        'last_followup_user_id': followup.user_id if followup else 'N/A',
                        'last_followup_date': followup.created_at.isoformat() if followup and followup.created_at else None,
                        'next_followup_date': followup.next_followup_date.isoformat() if followup and followup.next_followup_date else None,
                        'latest_flag': followup.followup_flag if followup else 'N/A',
                        'latest_remark': followup.followup_remark if followup else 'N/A',
                        'quotation_sent': followup.quotation_sent if followup else False,
                        'quotation_value': followup.quotation_value if followup else 0,
                        'quotation_no': followup.quotation_no if followup else None,
                        'source': 'followup_only'
                    }
                
                customers_list.append(customer_data)
            
            # ========== STEP 5: Sort customers by name and add serial numbers ==========
            customers_list.sort(key=lambda x: x['customer_name'])
            
            for idx, customer in enumerate(customers_list, 1):
                customer['s_no'] = idx
            
            # ========== STEP 6: Get follow-up summary stats ==========
            all_instance_ids_list = list(asset_numbers_set) + list(additional_instance_ids)
            
            total_followups = 0
            completed_followups = 0
            wip_followups = 0
            rejected_followups = 0
            rescheduled_followups = 0
            pending_followups = 0
            
            if all_instance_ids_list:
                # Single aggregate query instead of 6 separate count queries
                counts = db.query(
                    func.count(FollowUp.id).label('total'),
                    func.sum(case((FollowUp.status == 'completed', 1), else_=0)).label('completed'),
                    func.sum(case((FollowUp.status == 'wip', 1), else_=0)).label('wip'),
                    func.sum(case((FollowUp.status == 'rejected', 1), else_=0)).label('rejected'),
                    func.sum(case((FollowUp.status == 'rescheduled', 1), else_=0)).label('rescheduled'),
                    func.sum(case((FollowUp.status == 'pending', 1), else_=0)).label('pending'),
                ).filter(
                    FollowUp.campaign_id == campaign_id,
                    FollowUp.customer_instance_id.in_(all_instance_ids_list)
                ).first()

                total_followups       = int(counts.total or 0)
                completed_followups   = int(counts.completed or 0)
                wip_followups         = int(counts.wip or 0)
                rejected_followups    = int(counts.rejected or 0)
                rescheduled_followups = int(counts.rescheduled or 0)
                pending_followups     = int(counts.pending or 0)
            else:
                total_followups = 0
                completed_followups = 0
                wip_followups = 0
                rejected_followups = 0
                rescheduled_followups = 0
                pending_followups = 0
            
            completion_rate = round((completed_followups / total_followups * 100) if total_followups > 0 else 0, 2)
            
            return {
                'campaign_id': campaign.id,
                'campaign_name': campaign.name,
                'service': campaign.service,
                'description': campaign.description,
                'status': campaign.status,
                'total_customers': len(all_instance_ids_list),
                'asset_numbers_count': len(asset_numbers),
                'additional_customers_count': len(additional_instance_ids),
                'total_followups': total_followups,
                'completed_followups': completed_followups,
                'wip_followups': wip_followups,
                'rejected_followups': rejected_followups,
                'rescheduled_followups': rescheduled_followups,
                'pending_followups': pending_followups,
                'completion_rate': completion_rate,
                'customers': customers_list
            }
            
        except Exception as e:
            print(f"Error in get_campaign_customers: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    @staticmethod
    async def get_branch_campaign_customers(db: Session, branch_code: str, user_info: dict = None):
        """
        Get all campaigns and their customers for a specific branch
        Shows ALL campaigns (even those with no follow-ups)
        """
        try:
            from app.models.campaign_model import Campaign
            from app.models.engagement_model import FollowUp
            from app.models.customer_model import Customer
            
            # Get ALL campaigns (both active and inactive)
            all_campaigns = db.query(Campaign).all()
            
            # Get all customers in this branch
            branch_customers = db.query(Customer).filter(
                Customer.branch_id == branch_code
            ).all()
            
            if not branch_customers:
                return {
                    'branch_code': branch_code,
                    'branch_name': get_branch_display_name(branch_code),
                    'total_campaigns': len(all_campaigns),
                    'total_customers': 0,
                    'total_followups': 0,
                    'completed_followups': 0,
                    'wip_followups': 0,
                    'rejected_followups': 0,
                    'rescheduled_followups': 0,
                    'not_connected_followups': 0,
                    'pending_followups': 0,
                    'branch_completion_rate': 0,
                    'campaigns': []
                }
            
            branch_customer_ids = [c.instance_id for c in branch_customers]
            customer_map = {c.instance_id: c for c in branch_customers}
            
            # Get all follow-ups for this branch's customers
            batch_size = 500
            all_followups = []
            
            for i in range(0, len(branch_customer_ids), batch_size):
                batch_ids = branch_customer_ids[i:i + batch_size]
                batch_followups = db.query(FollowUp).filter(
                    FollowUp.customer_instance_id.in_(batch_ids)
                ).all()
                all_followups.extend(batch_followups)
            
            # Group follow-ups by campaign
            followups_by_campaign = {}
            for fu in all_followups:
                if fu.campaign_id not in followups_by_campaign:
                    followups_by_campaign[fu.campaign_id] = []
                followups_by_campaign[fu.campaign_id].append(fu)
            
            campaign_data = []
            total_branch_customers = 0
            total_branch_followups = 0
            total_branch_completed = 0
            total_branch_wip = 0
            total_branch_rejected = 0
            total_branch_rescheduled = 0
            total_branch_not_connected = 0
            total_branch_pending = 0
            
            # Process ALL campaigns (even those with no follow-ups)
            for campaign in all_campaigns:
                campaign_followups = followups_by_campaign.get(campaign.id, [])
                
                # Get customers with follow-ups for this campaign
                customers_with_followups = set()
                for fu in campaign_followups:
                    if fu.customer_instance_id:
                        customers_with_followups.add(fu.customer_instance_id)
                
                # Get customers from asset_numbers that are in this branch
                asset_numbers = campaign.asset_numbers or []
                customers_from_assets = set()
                for asset_id in asset_numbers:
                    if asset_id in customer_map:
                        customers_from_assets.add(asset_id)
                
                # IMPORTANT: Total Allocate = Customers from assets + Customers with follow-ups (even if not in assets)
                # This is the key fix - Union of both sets
                total_allocate_customers = customers_from_assets.union(customers_with_followups)
                
                # Filter customer map to only include customers in total_allocate
                filtered_customer_map = {cid: customer_map[cid] for cid in total_allocate_customers if cid in customer_map}
                
                # Group follow-ups by customer for this campaign
                followups_by_customer = {}
                for fu in campaign_followups:
                    customer_id = fu.customer_instance_id
                    if customer_id:
                        if customer_id not in followups_by_customer:
                            followups_by_customer[customer_id] = []
                        followups_by_customer[customer_id].append(fu)
                
                customers_list = []
                
                # Process customers with follow-ups first (Engaged customers - keep this logic exactly as before)
                for customer_instance_id, followups_list in followups_by_customer.items():
                    customer = filtered_customer_map.get(customer_instance_id)
                    if not customer:
                        continue
                    
                    latest_followup = max(followups_list, key=lambda x: x.created_at if x.created_at else x.id)
                    
                    total_customer_followups = len(followups_list)
                    completed_count = len([fu for fu in followups_list if fu.status == 'completed'])
                    wip_count = len([fu for fu in followups_list if fu.status == 'wip'])
                    rejected_count = len([fu for fu in followups_list if fu.status == 'rejected'])
                    rescheduled_count = len([fu for fu in followups_list if fu.status == 'rescheduled'])
                    pending_count = len([fu for fu in followups_list if fu.status == 'remaining'])
                    
                    customer_data = {
                        's_no': 0,
                        'instance_id': customer.instance_id,
                        'customer_name': customer.customer_name or 'N/A',
                        'phone_number': customer.phone_number or 'N/A',
                        'email': customer.email or 'N/A',
                        'branch_id': customer.branch_id or 'N/A',
                        'location': customer.location or 'N/A',
                        'last_status': latest_followup.status if latest_followup else '-',
                        'csp_subtype': latest_followup.csp_subtype if latest_followup else None,
                        'last_followup_user_name': latest_followup.user_name if latest_followup else 'N/A',
                        'last_followup_user_id': latest_followup.user_id if latest_followup else 'N/A',
                        'last_followup_date': latest_followup.created_at.isoformat() if latest_followup and latest_followup.created_at else None,
                        'next_followup_date': latest_followup.next_followup_date.isoformat() if latest_followup and latest_followup.next_followup_date else None,
                        'latest_flag': latest_followup.followup_flag if latest_followup else 'N/A',
                        'latest_remark': latest_followup.followup_remark if latest_followup else 'N/A',
                        'quotation_sent': latest_followup.quotation_sent if latest_followup else False,
                        'quotation_value': latest_followup.quotation_value if latest_followup else 0,
                        'total_followups': total_customer_followups,
                        'completed_followups': completed_count,
                        'wip_followups': wip_count,
                        'rejected_followups': rejected_count,
                        'rescheduled_followups': rescheduled_count,
                        'pending_followups': pending_count
                    }
                    
                    customers_list.append(customer_data)
                
                # Process customers from asset_numbers that have NO follow-ups (Remaining customers)
                # These are customers who are in asset_numbers but have no follow-ups yet
                customers_without_followups = customers_from_assets - customers_with_followups
                for customer_instance_id in customers_without_followups:
                    customer = filtered_customer_map.get(customer_instance_id)
                    if not customer:
                        continue
                    
                    customer_data = {
                        's_no': 0,
                        'instance_id': customer.instance_id,
                        'customer_name': customer.customer_name or 'N/A',
                        'phone_number': customer.phone_number or 'N/A',
                        'email': customer.email or 'N/A',
                        'branch_id': customer.branch_id or 'N/A',
                        'location': customer.location or 'N/A',
                        'last_status': '-',
                        'csp_subtype': None,
                        'last_followup_user_name': 'N/A',
                        'last_followup_user_id': 'N/A',
                        'last_followup_date': None,
                        'next_followup_date': None,
                        'latest_flag': 'N/A',
                        'latest_remark': 'N/A',
                        'quotation_sent': False,
                        'quotation_value': 0,
                        'total_followups': 0,
                        'completed_followups': 0,
                        'wip_followups': 0,
                        'rejected_followups': 0,
                        'rescheduled_followups': 0,
                        'pending_followups': 0
                    }
                    
                    customers_list.append(customer_data)
                
                # Sort customers by name
                customers_list.sort(key=lambda x: x['customer_name'])
                
                for idx, cust in enumerate(customers_list, 1):
                    cust['s_no'] = idx
                
                # Calculate campaign stats based on latest follow-up status (ENGAGED customers only)
                # This logic remains EXACTLY the same as before
                completed_followups = len([c for c in customers_list if c['last_status'] == 'completed'])
                wip_followups = len([c for c in customers_list if c['last_status'] == 'wip'])
                rejected_followups = len([c for c in customers_list if c['last_status'] == 'rejected'])
                rescheduled_followups = len([c for c in customers_list if c['last_status'] == 'rescheduled'])
                not_connected_followups = len([c for c in customers_list if c['last_status'] == 'not_connected'])
                pending_followups = len([c for c in customers_list if c['last_status'] == 'pending'])
                
                # Engaged customers = customers with any follow-up (not pending status from assets without followups)
                engaged_customers = len(customers_with_followups)
                
                # Total allocate = ALL customers (engaged + remaining from assets)
                total_allocate = len(total_allocate_customers)
                
                # Calculate completion rate based on engaged customers only (keep same logic)
                completion_rate = round(
                    (completed_followups / engaged_customers * 100) if engaged_customers > 0 else 0, 2
                )
                
                campaign_data.append({
                    'campaign_id': campaign.id,
                    'campaign_name': campaign.name,
                    'service': campaign.service,
                    'description': campaign.description or '',
                    'status': campaign.status,
                    'total_customers': engaged_customers,  # This is ENGAGED customers (with follow-ups) - KEEP SAME
                    'total_allocate': total_allocate,  # NEW FIELD: Total assets (from assets + followups not in assets)
                    'total_followups': len(campaign_followups),
                    'completed_followups': completed_followups,
                    'wip_followups': wip_followups,
                    'rejected_followups': rejected_followups,
                    'rescheduled_followups': rescheduled_followups,
                    'not_connected_followups': not_connected_followups,
                    'pending_followups': pending_followups,
                    'completion_rate': completion_rate,
                    'customers': customers_list
                })
                
                total_branch_customers += engaged_customers  # Keep same - total engaged
                total_branch_followups += len(campaign_followups)
                total_branch_completed += completed_followups
                total_branch_wip += wip_followups
                total_branch_rejected += rejected_followups
                total_branch_rescheduled += rescheduled_followups
                total_branch_not_connected += not_connected_followups
                total_branch_pending += pending_followups
            
            # Sort campaigns by name
            campaign_data.sort(key=lambda x: x['campaign_name'])
            
            # Calculate branch completion rate based on engaged customers (keep same)
            branch_completion_rate = round(
                (total_branch_completed / total_branch_customers * 100)
                if total_branch_customers > 0 else 0, 2
            )
            
            # Calculate total branch allocate (sum of all campaign total_allocate)
            total_branch_allocate = sum([c['total_allocate'] for c in campaign_data])
            
            return {
                'branch_code': branch_code,
                'branch_name': get_branch_display_name(branch_code),
                'total_campaigns': len(all_campaigns),
                'total_customers': total_branch_customers,  # Total engaged customers (keep same)
                'total_allocate': total_branch_allocate,  # NEW: Total assets across all campaigns
                'total_followups': total_branch_followups,
                'completed_followups': total_branch_completed,
                'wip_followups': total_branch_wip,
                'rejected_followups': total_branch_rejected,
                'rescheduled_followups': total_branch_rescheduled,
                'not_connected_followups': total_branch_not_connected,
                'pending_followups': total_branch_pending,
                'branch_completion_rate': branch_completion_rate,
                'campaigns': campaign_data
            }
            
        except Exception as e:
            print(f"Error in get_branch_campaign_customers: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    @staticmethod
    async def get_branch_total_customers_count(db: Session, branch_code: str, user_info: dict = None):
        """
        Get remaining customers count for ALL campaigns (both active and inactive) in a branch
        Remaining customers = Customers in branch with instance_id in campaign.asset_numbers
        """
        try:
            from app.models.campaign_model import Campaign
            from app.models.customer_model import Customer
    
            campaigns = db.query(Campaign).all()
    
            branch_rows = db.query(Customer.instance_id).filter(
                Customer.branch_id == branch_code
            ).all()
            branch_set = {r[0] for r in branch_rows if r[0]}
    
            campaign_counts = []
            total_customers_set = set()
    
            for campaign in campaigns:
                asset_set = set(campaign.asset_numbers or [])
                if not asset_set:
                    continue
                matched = branch_set & asset_set   # same result as the IN() query
                if not matched:
                    continue
                total_customers_set.update(matched)
                campaign_counts.append({
                    "campaign_id": campaign.id,
                    "campaign_name": campaign.name,
                    "remaining_customers": len(matched),
                    "status": campaign.status
                })
    
            return {
                "branch_code": branch_code,
                "total_customers_across_campaigns": len(total_customers_set),
                "campaigns": campaign_counts
            }
    
        except Exception as e:
            print(f"Error in get_branch_total_customers_count: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    @staticmethod
    async def get_branch_customer_allocate_summary(db: Session, branch_code: str, user_info: dict = None):
        try:
            from app.models.campaign_model import Campaign
            from app.models.customer_model import Customer
            from app.models.engagement_model import FollowUp
    
            # Step 1: asset instance IDs across all campaigns (in-memory set)
            all_campaigns = db.query(Campaign).all()
            asset_instance_ids = set()
            for campaign in all_campaigns:
                asset_instance_ids.update(campaign.asset_numbers or [])
    
            # Step 2: ALL customers in this branch (one indexed query)
            branch_rows = db.query(Customer.instance_id).filter(
                Customer.branch_id == branch_code
            ).all()
            branch_instance_ids = {r[0] for r in branch_rows if r[0]}
    
            if not branch_instance_ids:
                return {
                    'branch_code': branch_code,
                    'branch_name': get_branch_display_name(branch_code),
                    'total_allocated_customers': 0,
                    'attended_customers': 0,
                    'attended_percentage': 0
                }
    
            # Step 3: branch customers who have at least one follow-up
            # ONE join query — DB does the matching, no chunked IN() round-trips
            attended_rows = db.query(
                func.distinct(FollowUp.customer_instance_id)
            ).join(
                Customer, Customer.instance_id == FollowUp.customer_instance_id
            ).filter(
                Customer.branch_id == branch_code
            ).all()
            attended_set = {r[0] for r in attended_rows if r[0]}
    
            # Step 4: allocated = branch customers in assets OR with a follow-up
            # (every attended customer is allocated by definition)
            allocated_ids = {
                iid for iid in branch_instance_ids if iid in asset_instance_ids
            } | attended_set
    
            total_allocated = len(allocated_ids)
            attended_customers = len(attended_set)
            attended_percentage = round(
                (attended_customers / total_allocated * 100), 2
            ) if total_allocated > 0 else 0
    
            return {
                'branch_code': branch_code,
                'branch_name': get_branch_display_name(branch_code),
                'total_allocated_customers': total_allocated,
                'attended_customers': attended_customers,
                'attended_percentage': attended_percentage
            }
    
        except Exception as e:
            print(f"Error in get_branch_customer_allocate_summary: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    @staticmethod
    async def get_all_branches_campaign_summary(db: Session, only_branch: str = None):
        """
        ONE-call replacement for the dashboard's per-branch fan-out over
        branch-campaign-customers + branch-total-customers + allocate-summary.

        Returns, per branch, the SAME aggregate numbers those endpoints produce —
        shaped exactly like their responses minus the heavy per-customer lists
        (the dashboard charts read only aggregates; the Branch Customers modal
        still calls the per-branch endpoints when it is actually opened).

        Data is pulled with three slim queries (campaigns, 2-column customer map,
        5-column follow-ups) instead of N branches × full ORM loads.
        """
        try:
            from app.models.campaign_model import Campaign
            from app.models.customer_model import Customer
            from app.models.engagement_model import FollowUp

            campaigns = db.query(Campaign).all()

            # instance_id -> branch for every customer (two indexed columns)
            cust_branch = {}
            branch_customers = {}
            for iid, br in db.query(Customer.instance_id, Customer.branch_id).all():
                if not iid:
                    continue
                cust_branch[iid] = br
                if br:
                    branch_customers.setdefault(br, set()).add(iid)

            # Also seed every branch that has USERS (e.g. HO) even when no customer
            # is assigned to it — otherwise those branches are missing from the
            # response and their dashboard box never leaves the loading skeleton.
            from app.models.user_model import User as _User
            for (br,) in db.query(_User.branch).distinct().all():
                if br:
                    branch_customers.setdefault(br, set())

            if only_branch is not None:
                branch_customers = {only_branch: branch_customers.get(only_branch, set())}

            # Latest follow-up per (campaign, customer) + per-pair count — slim columns.
            # Mirrors the per-branch logic: campaign stats are based on each engaged
            # customer's LATEST follow-up status within that campaign.
            DT_MIN = datetime(1900, 1, 1)
            latest = {}           # (campaign_id, iid) -> [sort_key, status, count]
            branch_attended = {}  # branch -> set(iid) with ANY follow-up (for allocation)
            for cid, iid, status_val, created, fid in db.query(
                FollowUp.campaign_id,
                FollowUp.customer_instance_id,
                FollowUp.status,
                FollowUp.created_at,
                FollowUp.id,
            ).all():
                if not iid:
                    continue
                br = cust_branch.get(iid)
                if br is None or br not in branch_customers:
                    continue
                branch_attended.setdefault(br, set()).add(iid)
                key = (cid, iid)
                sort_key = (created or DT_MIN, fid or 0)
                cur = latest.get(key)
                if cur is None:
                    latest[key] = [sort_key, status_val, 1]
                else:
                    cur[2] += 1
                    if sort_key > cur[0]:
                        cur[0] = sort_key
                        cur[1] = status_val

            # Aggregate engaged stats per (campaign, branch)
            camp_branch = {}
            for (cid, iid), (skey, st, cnt) in latest.items():
                br = cust_branch.get(iid)
                agg = camp_branch.setdefault((cid, br), {
                    'engaged_ids': set(), 'followups': 0,
                    'completed': 0, 'wip': 0, 'rejected': 0,
                    'rescheduled': 0, 'not_connected': 0, 'pending': 0,
                })
                agg['engaged_ids'].add(iid)
                agg['followups'] += cnt
                s = (st or '').lower()
                if s in ('completed', 'wip', 'rejected', 'rescheduled', 'not_connected', 'pending'):
                    agg[s] += 1

            all_asset_ids = set()
            campaign_assets = {}
            for campaign in campaigns:
                asset_set = set(campaign.asset_numbers or [])
                campaign_assets[campaign.id] = asset_set
                all_asset_ids |= asset_set

            result = {}
            for br, br_cust_set in branch_customers.items():
                campaign_list = []
                remaining_campaigns = []
                remaining_union = set()
                totals = {'engaged': 0, 'followups': 0, 'completed': 0, 'wip': 0,
                          'rejected': 0, 'rescheduled': 0, 'not_connected': 0, 'pending': 0}
                total_allocate_sum = 0

                for campaign in campaigns:
                    agg = camp_branch.get((campaign.id, br))
                    engaged_ids = agg['engaged_ids'] if agg else set()
                    matched_assets = campaign_assets[campaign.id] & br_cust_set
                    total_allocate = len(matched_assets | engaged_ids)
                    engaged = len(engaged_ids)
                    completed = agg['completed'] if agg else 0

                    campaign_list.append({
                        'campaign_id': campaign.id,
                        'campaign_name': campaign.name,
                        'service': campaign.service,
                        'description': campaign.description or '',
                        'status': campaign.status,
                        'total_customers': engaged,
                        'total_allocate': total_allocate,
                        'total_followups': agg['followups'] if agg else 0,
                        'completed_followups': completed,
                        'wip_followups': agg['wip'] if agg else 0,
                        'rejected_followups': agg['rejected'] if agg else 0,
                        'rescheduled_followups': agg['rescheduled'] if agg else 0,
                        'not_connected_followups': agg['not_connected'] if agg else 0,
                        'pending_followups': agg['pending'] if agg else 0,
                        'completion_rate': round((completed / engaged * 100) if engaged > 0 else 0, 2),
                    })

                    if matched_assets:
                        remaining_union |= matched_assets
                        remaining_campaigns.append({
                            'campaign_id': campaign.id,
                            'campaign_name': campaign.name,
                            'remaining_customers': len(matched_assets),
                            'status': campaign.status,
                        })

                    totals['engaged'] += engaged
                    totals['followups'] += agg['followups'] if agg else 0
                    for k in ('completed', 'wip', 'rejected', 'rescheduled', 'not_connected', 'pending'):
                        totals[k] += agg[k] if agg else 0
                    total_allocate_sum += total_allocate

                campaign_list.sort(key=lambda x: x['campaign_name'])

                attended_set = branch_attended.get(br, set())
                allocated = (br_cust_set & all_asset_ids) | attended_set

                result[br] = {
                    'engaged': {
                        'branch_code': br,
                        'branch_name': get_branch_display_name(br),
                        'total_campaigns': len(campaigns),
                        'total_customers': totals['engaged'],
                        'total_allocate': total_allocate_sum,
                        'total_followups': totals['followups'],
                        'completed_followups': totals['completed'],
                        'wip_followups': totals['wip'],
                        'rejected_followups': totals['rejected'],
                        'rescheduled_followups': totals['rescheduled'],
                        'not_connected_followups': totals['not_connected'],
                        'pending_followups': totals['pending'],
                        'branch_completion_rate': round(
                            (totals['completed'] / totals['engaged'] * 100)
                            if totals['engaged'] > 0 else 0, 2),
                        'campaigns': campaign_list,
                    },
                    'remaining': {
                        'branch_code': br,
                        'total_customers_across_campaigns': len(remaining_union),
                        'campaigns': remaining_campaigns,
                    },
                    'allocation': {
                        'branch_code': br,
                        'branch_name': get_branch_display_name(br),
                        'total_allocated_customers': len(allocated),
                        'attended_customers': len(attended_set),
                        'attended_percentage': round(
                            (len(attended_set) / len(allocated) * 100)
                            if allocated else 0, 2),
                    },
                }

            return result

        except Exception as e:
            print(f"Error in get_all_branches_campaign_summary: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    @staticmethod
    async def get_non_followup_unique_customer_stats(db: Session, user_id: str):
        """
        Get unique customer count and their latest status breakdown
        from non_followups table for a specific user
        """
        try:
            from app.models.non_followup_model import NonFollowUp
    
            all_non_followups = db.query(NonFollowUp).filter(
                NonFollowUp.user_id == user_id,
                NonFollowUp.customer_instance_id.isnot(None)
            ).all()
    
            # Track latest non-followup per unique customer_instance_id
            latest_map = {}
            for nfu in all_non_followups:
                instance_id = nfu.customer_instance_id
                record_date = nfu.followup_date if nfu.followup_date else nfu.created_at
    
                if instance_id not in latest_map:
                    latest_map[instance_id] = nfu
                else:
                    existing_date = (
                        latest_map[instance_id].followup_date
                        if latest_map[instance_id].followup_date
                        else latest_map[instance_id].created_at
                    )
                    if record_date and existing_date and record_date > existing_date:
                        latest_map[instance_id] = nfu
    
            total_unique = len(latest_map)
            completed = 0
            wip = 0
            rejected = 0
            rescheduled = 0
            not_connected = 0
            pending = 0
    
            for instance_id, nfu in latest_map.items():
                status = (nfu.status or '').lower()
                if status == 'completed':
                    completed += 1
                elif status == 'wip':
                    wip += 1
                elif status == 'rejected':
                    rejected += 1
                elif status == 'rescheduled':
                    rescheduled += 1
                elif status == 'not_connected':
                    not_connected += 1
                else:
                    pending += 1
    
            return {
                "total_unique_customers": total_unique,
                "completed": completed,
                "wip": wip,
                "rejected": rejected,
                "rescheduled": rescheduled,
                "not_connected": not_connected,
                "pending": pending
            }
    
        except Exception as e:
            print(f"Error in get_non_followup_unique_customer_stats: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                "total_unique_customers": 0,
                "completed": 0,
                "wip": 0,
                "rejected": 0,
                "rescheduled": 0,
                "not_connected": 0,
                "pending": 0
            }            
        
    @staticmethod
    async def get_non_campaign_customers(db: Session, user_info: dict = None):
        """
        Get all customers from non_followups table with their latest follow-up details.
        Returns unique customers based on customer_instance_id — latest follow-up per customer.
        """
        try:
            from app.models.non_followup_model import NonFollowUp
            from app.models.customer_model import Customer
    
            all_non_followups = db.query(NonFollowUp).all()
    
            if not all_non_followups:
                return {"total_customers": 0, "customers": []}
    
            # Build latest followup map per customer_instance_id
            latest_map = {}
            for nfu in all_non_followups:
                instance_id = nfu.customer_instance_id
                if not instance_id:
                    continue
                record_date = nfu.followup_date if nfu.followup_date else nfu.created_at
                if instance_id not in latest_map:
                    latest_map[instance_id] = nfu
                else:
                    existing_date = (
                        latest_map[instance_id].followup_date
                        if latest_map[instance_id].followup_date
                        else latest_map[instance_id].created_at
                    )
                    if record_date and existing_date and record_date > existing_date:
                        latest_map[instance_id] = nfu
    
            # Fetch customer details in batch
            instance_ids = list(latest_map.keys())
            customers_db = db.query(Customer).filter(
                Customer.instance_id.in_(instance_ids)
            ).all() if instance_ids else []
            customer_map = {c.instance_id: c for c in customers_db}

            # Bulk-fetch Activity content for the latest non-followups (same as
            # get_my_non_campaign_customers). This was missing here — the loop
            # below referenced an undefined `activity`, so the endpoint always
            # errored into its except-handler and returned an empty list.
            from app.models.engagement_model import Activity
            activity_ids = list({nfu.activity_id for nfu in latest_map.values() if nfu.activity_id})
            activity_map = {}
            if activity_ids:
                for a in db.query(Activity).filter(Activity.id.in_(activity_ids)).all():
                    activity_map[a.id] = a

            customers_list = []
            for idx, (instance_id, nfu) in enumerate(latest_map.items(), 1):
                customer = customer_map.get(instance_id)
                activity = activity_map.get(nfu.activity_id) if nfu.activity_id else None
                customers_list.append({
                    "s_no": idx,
                    "instance_id": instance_id,
                    "customer_name": customer.customer_name if customer else "Not Found",
                    "phone_number": customer.phone_number if customer else "N/A",
                    "email": customer.email if customer else "N/A",
                    "branch_id": customer.branch_id if customer else "N/A",
                    "location": customer.location if customer else "N/A",
                    "service": nfu.service or "N/A",
                    "remark_type": nfu.remark_type or "other",
                    "activity_content": activity.content if activity else None,
                    "last_status": nfu.status,
                    "last_followup_user_name": nfu.user_name or "N/A",
                    "last_followup_user_id": nfu.user_id or "N/A",
                    "last_followup_date": nfu.followup_date.isoformat() if nfu.followup_date else (
                        nfu.created_at.isoformat() if nfu.created_at else None
                    ),
                    "next_followup_date": nfu.next_followup_date.isoformat() if nfu.next_followup_date else None,
                    "latest_flag": nfu.followup_flag or "N/A",
                    "latest_remark": nfu.followup_remark or "N/A",
                    "followup_by": nfu.followup_by or "N/A",
                    "quotation_sent": nfu.quotation_sent or False,
                    "quotation_value": float(nfu.quotation_value) if nfu.quotation_value else 0,
                    "quotation_no": nfu.quotation_no or None,
                })
    
            # Sort by last_followup_date descending
            customers_list.sort(
                key=lambda x: x["last_followup_date"] or "",
                reverse=True
            )
            for idx, c in enumerate(customers_list, 1):
                c["s_no"] = idx
    
            return {
                "total_customers": len(customers_list),
                "customers": customers_list
            }
    
        except Exception as e:
            print(f"Error in get_non_campaign_customers: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"total_customers": 0, "customers": []}        

    @staticmethod
    async def get_employee_campaign_progress(db: Session, employee_id: str, time_period: str = 'all', start_date=None, end_date=None):
        """
        Get per-campaign performance for a specific employee.
        Shows unique customer latest-followup status counts for each active campaign.

        Intentionally does NOT use apply_time_filter — filters on followup_date
        (the IST date the employee chose when logging) with full-day IST boundaries
        (00:00:00 → 23:59:59) so Today/Yesterday/custom ranges match exactly what
        the employee sees on their own screen.  created_at is used only as a
        fallback when followup_date is NULL.
        """
        try:
            from app.models.campaign_model import Campaign
            from app.models.engagement_model import FollowUp

            active_campaigns = (
                db.query(Campaign)
                .filter(Campaign.status == 'active')
                .order_by(Campaign.name)
                .all()
            )

            # ── Build IST-aware date boundaries ──────────────────────────────
            # followup_date is stored as an IST DateTime (SQL Server GETDATE()).
            # We want the full calendar day in IST, so:
            #   start_dt = YYYY-MM-DD 00:00:00  (inclusive)
            #   end_dt   = YYYY-MM-DD 23:59:59.999999  (inclusive)
            # Both as IST-naive datetimes (no tzinfo) to match the column type.

            start_dt: datetime | None = None
            end_dt:   datetime | None = None

            if time_period == 'all' or not time_period:
                pass  # no date filter — return all records

            elif time_period == 'custom':
                # start_date / end_date arrive as 'YYYY-MM-DD' strings from the route
                if start_date and end_date:
                    try:
                        sd = datetime.strptime(str(start_date)[:10], '%Y-%m-%d')
                        ed = datetime.strptime(str(end_date)[:10],   '%Y-%m-%d')
                        start_dt = sd.replace(hour=0,  minute=0,  second=0,  microsecond=0)
                        end_dt   = ed.replace(hour=23, minute=59, second=59, microsecond=999999)
                    except Exception as parse_err:
                        print(f"[employee_campaign_progress] date parse error: {parse_err}")

            else:
                # Safety-net for any named periods the caller might send directly
                # (frontend normally converts these to 'custom' + explicit dates).
                now_ist = datetime.now(IST).replace(tzinfo=None)
                today_start = now_ist.replace(hour=0, minute=0, second=0, microsecond=0)
                today_end   = now_ist.replace(hour=23, minute=59, second=59, microsecond=999999)

                if time_period == 'today':
                    start_dt = today_start
                    end_dt   = today_end
                elif time_period == 'yesterday':
                    yesterday = today_start - timedelta(days=1)
                    start_dt  = yesterday
                    end_dt    = yesterday.replace(hour=23, minute=59, second=59, microsecond=999999)
                elif time_period == 'week':
                    start_dt = today_start - timedelta(days=6)
                    end_dt   = today_end
                elif time_period == 'month':
                    start_dt = today_start - timedelta(days=29)
                    end_dt   = today_end
                # any other unknown period → no filter (treat as 'all')

            # ── Batched fetch (fixes N+1: was one query per active campaign) ──
            # Same filters as before, but ONE query over all active campaign ids
            # (chunked IN() for MSSQL's parameter cap), grouped per campaign.
            followups_by_campaign: dict = {}
            active_campaign_ids = [c.id for c in active_campaigns]
            CHUNK = 1000
            for i in range(0, len(active_campaign_ids), CHUNK):
                chunk = active_campaign_ids[i:i + CHUNK]
                base_q = db.query(FollowUp).filter(
                    FollowUp.campaign_id.in_(chunk),
                    FollowUp.user_id == employee_id,
                    FollowUp.customer_instance_id.isnot(None),
                )

                if start_dt and end_dt:
                    # Primary: filter on followup_date (the date the employee chose).
                    # Fallback: when followup_date is NULL, fall back to created_at
                    # so records saved without an explicit date are still included.
                    base_q = base_q.filter(
                        or_(
                            and_(
                                FollowUp.followup_date.isnot(None),
                                FollowUp.followup_date >= start_dt,
                                FollowUp.followup_date <= end_dt,
                            ),
                            and_(
                                FollowUp.followup_date.is_(None),
                                FollowUp.created_at >= start_dt,
                                FollowUp.created_at <= end_dt,
                            ),
                        )
                    )

                for fu in base_q.all():
                    followups_by_campaign.setdefault(fu.campaign_id, []).append(fu)

            # ── Per-campaign aggregation (unchanged logic) ────────────────────
            result = []

            for campaign in active_campaigns:
                all_followups = followups_by_campaign.get(campaign.id, [])

                if not all_followups:
                    continue

                # ── Latest followup per unique customer ───────────────────────
                latest_map: dict = {}
                for fu in all_followups:
                    iid     = fu.customer_instance_id
                    fu_date = fu.followup_date if fu.followup_date else fu.created_at
                    if iid not in latest_map:
                        latest_map[iid] = fu
                    else:
                        existing_date = (
                            latest_map[iid].followup_date
                            if latest_map[iid].followup_date
                            else latest_map[iid].created_at
                        )
                        if fu_date and existing_date and fu_date > existing_date:
                            latest_map[iid] = fu

                total_unique  = len(latest_map)
                completed     = sum(1 for fu in latest_map.values() if fu.status == 'completed')
                wip           = sum(1 for fu in latest_map.values() if fu.status == 'wip')
                rescheduled   = sum(1 for fu in latest_map.values() if fu.status == 'rescheduled')
                rejected      = sum(1 for fu in latest_map.values() if fu.status == 'rejected')
                not_connected = sum(1 for fu in latest_map.values() if fu.status == 'not_connected')
                pending       = sum(1 for fu in latest_map.values() if fu.status == 'pending')

                total_followups = len(all_followups)
                success_pct     = round((completed / total_unique * 100) if total_unique > 0 else 0, 1)

                result.append({
                    'campaign_id':            campaign.id,
                    'campaign_name':          campaign.name,
                    'service':                campaign.service,
                    'status':                 campaign.status,
                    'total_unique_customers': total_unique,
                    'total_followups':        total_followups,
                    'completed':              completed,
                    'wip':                    wip,
                    'rescheduled':            rescheduled,
                    'rejected':               rejected,
                    'not_connected':          not_connected,
                    'pending':                pending,
                    'success_pct':            success_pct,
                })

            result.sort(key=lambda x: x['total_unique_customers'], reverse=True)
            return result

        except Exception as e:
            print(f"Error in get_employee_campaign_progress: {str(e)}")
            import traceback
            traceback.print_exc()
            return []            
        
    @staticmethod
    async def get_user_all_followups(db: Session, user_id: str, time_period: str = 'all', start_date=None, end_date=None):
        """
        Get ALL follow-ups taken by a specific user with all columns from followup table.
        Includes customer details and campaign name for display.
        """
        try:
            from app.models.engagement_model import FollowUp, Activity, RR
            from app.models.campaign_model import Campaign
            from app.models.customer_model import Customer
    
            base_query = db.query(FollowUp).filter(FollowUp.user_id == user_id)
            filtered_query = EmployeePerformanceController.apply_time_filter(
                base_query, FollowUp, time_period, start_date, end_date
            )
    
            followups = filtered_query.order_by(FollowUp.followup_date.desc()).all()
    
            if not followups:
                return {"total": 0, "followups": []}
    
            # Bulk-fetch related data
            campaign_ids = list({fu.campaign_id for fu in followups if fu.campaign_id})
            customer_ids = list({fu.customer_id for fu in followups if fu.customer_id})
            activity_ids = list({fu.activity_id for fu in followups if fu.activity_id})
            rr_ids = list({fu.rr_id for fu in followups if fu.rr_id})
    
            campaigns_map = {}
            if campaign_ids:
                for c in db.query(Campaign).filter(Campaign.id.in_(campaign_ids)).all():
                    campaigns_map[c.id] = c
    
            customers_map = {}
            if customer_ids:
                for c in db.query(Customer).filter(Customer.id.in_(customer_ids)).all():
                    customers_map[c.id] = c
    
            activities_map = {}
            if activity_ids:
                for a in db.query(Activity).filter(Activity.id.in_(activity_ids)).all():
                    activities_map[a.id] = a
    
            rr_map = {}
            if rr_ids:
                for r in db.query(RR).filter(RR.id.in_(rr_ids)).all():
                    rr_map[r.id] = r
    
            result = []
            for idx, fu in enumerate(followups, 1):
                campaign = campaigns_map.get(fu.campaign_id)
                customer = customers_map.get(fu.customer_id)
                activity = activities_map.get(fu.activity_id) if fu.activity_id else None
                rr = rr_map.get(fu.rr_id) if fu.rr_id else None
    
                result.append({
                    "s_no": idx,
                    "id": fu.id,
                    "customer_id": fu.customer_id,
                    "customer_instance_id": fu.customer_instance_id,
                    "customer_name": customer.customer_name if customer else 'N/A',
                    "phone_number": customer.phone_number if customer else 'N/A',
                    "email": customer.email if customer else 'N/A',
                    "branch_id": customer.branch_id if customer else 'N/A',
                    "campaign_id": fu.campaign_id,
                    "campaign_name": campaign.name if campaign else 'N/A',
                    "campaign_service": campaign.service if campaign else 'N/A',
                    "user_id": fu.user_id,
                    "user_name": fu.user_name,
                    "followup_date": fu.followup_date.isoformat() if fu.followup_date else None,
                    "followup_by": fu.followup_by,
                    "csp_subtype": fu.csp_subtype,
                    "followup_flag": fu.followup_flag,
                    "followup_remark": fu.followup_remark,
                    "status": fu.status,
                    "next_followup_date": fu.next_followup_date.isoformat() if fu.next_followup_date else None,
                    "quotation_sent": fu.quotation_sent or False,
                    "quotation_no": fu.quotation_no,
                    "quotation_value": float(fu.quotation_value) if fu.quotation_value else 0,
                    "activity_id": fu.activity_id,
                    "activity_content": activity.content if activity else None,
                    "rr_id": fu.rr_id,
                    "rr_content": rr.content if rr else None,
                    "created_at": fu.created_at.isoformat() if fu.created_at else None,
                    "updated_at": fu.updated_at.isoformat() if fu.updated_at else None,
                })
    
            return {"total": len(result), "followups": result}
    
        except Exception as e:
            print(f"Error in get_user_all_followups: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"total": 0, "followups": []}      

    @staticmethod
    async def get_user_letter_count(db: Session, user_id: str):
        """
        Cheap COUNT of letters sent BY this user (for the dashboard card),
        split by status so the card can show Sent / Draft separately.
        """
        try:
            from app.models.letter_model import LetterSendRecord
            row = db.query(
                func.count(LetterSendRecord.id).label('total'),
                func.sum(case((func.lower(LetterSendRecord.status) == 'sent', 1), else_=0)).label('sent'),
                func.sum(case((func.lower(LetterSendRecord.status) == 'draft', 1), else_=0)).label('draft'),
                # Letters whose Format Type name starts with "CSP" (e.g. warranty-lapse CSP letters)
                func.sum(case((func.lower(LetterSendRecord.format_type_name).like('csp%'), 1), else_=0)).label('csp'),
            ).filter(
                LetterSendRecord.sent_by_id == user_id
            ).first()
            return {
                "count": int(row.total or 0),
                "sent": int(row.sent or 0),
                "draft": int(row.draft or 0),
                "csp": int(row.csp or 0),
            }
        except Exception as e:
            print(f"Error in get_user_letter_count: {str(e)}")
            return {"count": 0, "sent": 0, "draft": 0, "csp": 0}

    @staticmethod
    async def get_user_letter_records(db: Session, user_id: str):
        """
        All letters sent BY a specific user.
        PERF: selects ONLY light columns — letter_html / letter_body / letter_fields /
        attachments are intentionally NOT fetched (they can be huge base64/HTML blobs).
        Customer name/phone/branch are pulled in ONE batched IN() query.
        """
        try:
            from app.models.letter_model import LetterSendRecord
            from app.models.customer_model import Customer

            # PERF: `attachments` stores full base64 file CONTENT — never select it
            # for a listing. Names are extracted inside SQL Server (OPENJSON 2016+,
            # STRING_AGG 2017+) with a fallback to the old column select.
            base_cols = [
                LetterSendRecord.id,
                LetterSendRecord.ref_no,
                LetterSendRecord.financial_year,
                LetterSendRecord.sequence_number,
                LetterSendRecord.instance_id,
                LetterSendRecord.customer_id,
                LetterSendRecord.format_type_name,
                LetterSendRecord.subject,
                LetterSendRecord.channels,
                LetterSendRecord.sent_email,
                LetterSendRecord.sent_whatsapp,
                LetterSendRecord.email_to,
                LetterSendRecord.email_cc,          # NEW — only if your model has this column (see note below)
                LetterSendRecord.whatsapp_to,
                LetterSendRecord.status,
                LetterSendRecord.error_message,
                LetterSendRecord.sent_by_id,
                LetterSendRecord.sent_by_name,
                LetterSendRecord.created_at,
            ]
            from sqlalchemy import literal_column
            names_expr = literal_column(
                "(SELECT STRING_AGG(COALESCE(JSON_VALUE(a.[value], '$.name'), "
                "CASE WHEN a.[type] = 1 THEN a.[value] END), '||') "
                "FROM OPENJSON(letter_send_records.attachments) a)"
            ).label('attachment_names')
            fast_names = True
            try:
                rows = db.query(*base_cols, names_expr).filter(
                    LetterSendRecord.sent_by_id == user_id
                ).order_by(LetterSendRecord.created_at.desc()).all()
            except Exception:
                db.rollback()
                fast_names = False
                rows = db.query(*base_cols, LetterSendRecord.attachments).filter(
                    LetterSendRecord.sent_by_id == user_id
                ).order_by(LetterSendRecord.created_at.desc()).all()

            if not rows:
                return {"total": 0, "letters": []}

            # Batch customer lookup (one indexed IN query)
            instance_ids = list({r.instance_id for r in rows if r.instance_id})
            customer_map = {}
            if instance_ids:
                for c in db.query(
                    Customer.instance_id, Customer.customer_name,
                    Customer.phone_number, Customer.branch_id
                ).filter(Customer.instance_id.in_(instance_ids)).all():
                    customer_map[c.instance_id] = c

            letters = []
            for idx, r in enumerate(rows, 1):
                cust = customer_map.get(r.instance_id)

                # File NAMES only — drop the heavy base64 `content` before sending to the UI
                if fast_names:
                    attachment_names = [n for n in (r.attachment_names or '').split('||') if n]
                else:
                    attachment_names = [
                        (a.get("name") if isinstance(a, dict) else a)
                        for a in (r.attachments or [])
                        if (a.get("name") if isinstance(a, dict) else a)
                    ]

                letters.append({
                    "s_no": idx,
                    "id": r.id,
                    "ref_no": r.ref_no or "-",
                    "financial_year": r.financial_year or "-",
                    "sequence_number": r.sequence_number,
                    "instance_id": r.instance_id or "-",
                    "customer_id": r.customer_id,
                    "customer_name": cust.customer_name if cust else "-",
                    "phone_number": cust.phone_number if cust else "-",
                    "branch_id": cust.branch_id if cust else "-",
                    "format_type_name": r.format_type_name or "-",
                    "subject": r.subject or "-",
                    "channels": r.channels or [],
                    "sent_email": bool(r.sent_email),
                    "sent_whatsapp": bool(r.sent_whatsapp),
                    "email_to": r.email_to or "-",
                    "email_cc": r.email_cc or "-",          # NEW — only if your model has email_cc
                    "whatsapp_to": r.whatsapp_to or "-",
                    "attachment_names": attachment_names,    # NEW
                    "status": r.status or "-",
                    "error_message": r.error_message or "",
                    "sent_by_name": r.sent_by_name or "-",
                    "created_at": r.created_at.isoformat() if r.created_at else None,
                })

            return {"total": len(letters), "letters": letters}

        except Exception as e:
            print(f"Error in get_user_letter_records: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"total": 0, "letters": []}

    @staticmethod
    async def get_user_letter_html(db: Session, user_id: str, letter_id: int):
        """
        Return ONE letter's stored HTML (+ ref_no/subject) so the FRONTEND renders it
        to a properly-formatted PDF (same letterhead as the Send Letter feature).
        Verifies the letter belongs to this user (blocks viewing someone else's letter
        by changing the id in the URL). Returns a dict or None.
        """
        try:
            from app.models.letter_model import LetterSendRecord   # same path as get_user_letter_records

            rec = db.query(
                LetterSendRecord.ref_no,
                LetterSendRecord.subject,
                LetterSendRecord.letter_html,
                LetterSendRecord.letter_body,
            ).filter(
                LetterSendRecord.id == letter_id,
                LetterSendRecord.sent_by_id == user_id          # ownership check
            ).first()

            if not rec:
                return None

            html = rec.letter_html
            if not html or not str(html).strip():
                body = (rec.letter_body or '').replace('\n', '<br/>')
                html = f"<div style='font-family:Arial;padding:24px;'><h3>{rec.subject or ''}</h3><div>{body}</div></div>"

            return {
                "ref_no": rec.ref_no or f"letter_{letter_id}",
                "subject": rec.subject or "",
                "letter_html": str(html),
            }

        except Exception as e:
            print(f"Error in get_user_letter_html: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    @staticmethod
    async def get_asset_lookup(db: Session, instance_id: str):
        """
        Lookup goem_oem and segment from asset_detailed by instance_id (asset number).
        """
        try:
            from app.models.customer_model import AssetDetailed
    
            asset = db.query(AssetDetailed).filter(
                (AssetDetailed.instance_id == instance_id) |
                (AssetDetailed.asset_number == instance_id)
            ).first()
    
            if not asset:
                return {"found": False, "goem_oem": "", "segment": "", "branch_id": ""}
    
            return {
                "found": True,
                "goem_oem": asset.goem_oem or "",
                "segment": asset.segment or "",
                "branch_id": asset.branch_id or ""
            }
        except Exception as e:
            print(f"Error in get_asset_lookup: {str(e)}")
            return {"found": False, "goem_oem": "", "segment": ""}       

    @staticmethod
    async def get_my_non_campaign_customers(db: Session, user_id: str):
        """
        Get all non-campaign customers for a SPECIFIC user with their latest follow-up
        details. Returns unique customers based on customer_instance_id (latest per customer).
        """
        try:
            from app.models.non_followup_model import NonFollowUp
            from app.models.customer_model import Customer

            all_non_followups = db.query(NonFollowUp).filter(
                NonFollowUp.user_id == user_id
            ).all()

            if not all_non_followups:
                return {"total_customers": 0, "customers": []}

            # Latest non-followup per customer_instance_id
            latest_map = {}
            for nfu in all_non_followups:
                instance_id = nfu.customer_instance_id
                if not instance_id:
                    continue
                record_date = nfu.followup_date if nfu.followup_date else nfu.created_at
                if instance_id not in latest_map:
                    latest_map[instance_id] = nfu
                else:
                    existing_date = (
                        latest_map[instance_id].followup_date
                        if latest_map[instance_id].followup_date
                        else latest_map[instance_id].created_at
                    )
                    if record_date and existing_date and record_date > existing_date:
                        latest_map[instance_id] = nfu

            instance_ids = list(latest_map.keys())
            customers_db = db.query(Customer).filter(
                Customer.instance_id.in_(instance_ids)
            ).all() if instance_ids else []
            customer_map = {c.instance_id: c for c in customers_db}

            # Bulk-fetch Activity content for the latest non-followups so the
            # All-Follow-ups modal can show an Activity for the "other" rows.
            from app.models.engagement_model import Activity
            activity_ids = list({nfu.activity_id for nfu in latest_map.values() if nfu.activity_id})
            activity_map = {}
            if activity_ids:
                for a in db.query(Activity).filter(Activity.id.in_(activity_ids)).all():
                    activity_map[a.id] = a

            customers_list = []
            for instance_id, nfu in latest_map.items():
                customer = customer_map.get(instance_id)
                activity = activity_map.get(nfu.activity_id) if nfu.activity_id else None
                customers_list.append({
                    "s_no": 0,
                    "instance_id": instance_id,
                    "customer_name": customer.customer_name if customer else "Not Found",
                    "phone_number": customer.phone_number if customer else "N/A",
                    "email": customer.email if customer else "N/A",
                    "branch_id": customer.branch_id if customer else "N/A",
                    "location": customer.location if customer else "N/A",
                    "service": nfu.service or "N/A",
                    "remark_type": nfu.remark_type or "other",
                    "activity_content": activity.content if activity else None,
                    "last_status": nfu.status,
                    "last_followup_user_name": nfu.user_name or "N/A",
                    "last_followup_user_id": nfu.user_id or "N/A",
                    "last_followup_date": nfu.followup_date.isoformat() if nfu.followup_date else (
                        nfu.created_at.isoformat() if nfu.created_at else None
                    ),
                    "next_followup_date": nfu.next_followup_date.isoformat() if nfu.next_followup_date else None,
                    "latest_flag": nfu.followup_flag or "N/A",
                    "latest_remark": nfu.followup_remark or "N/A",
                    "followup_by": nfu.followup_by or "N/A",
                    "quotation_sent": nfu.quotation_sent or False,
                    "quotation_value": float(nfu.quotation_value) if nfu.quotation_value else 0,
                    "quotation_no": nfu.quotation_no or None,
                })

            customers_list.sort(key=lambda x: x["last_followup_date"] or "", reverse=True)
            for idx, c in enumerate(customers_list, 1):
                c["s_no"] = idx

            return {"total_customers": len(customers_list), "customers": customers_list}

        except Exception as e:
            print(f"Error in get_my_non_campaign_customers: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"total_customers": 0, "customers": []}       
            
    @staticmethod
    async def get_employee_time_report(db: Session, target_date: str = None,
                                       branch_code: str = None, search: str = None):
        """
        Login / logout report (IST). Defaults to today's IST date.
        Work time is computed on the fly from login/logout — nothing daily is stored.
        """
        try:
            from datetime import timedelta
            from app.models.login_activity_model import LoginSession, now_ist
            from app.models.user_model import User

            # Resolve the date (IST). Default = today in IST.
            if target_date:
                try:
                    day = datetime.strptime(target_date, '%Y-%m-%d').date()
                except Exception:
                    day = now_ist().date()
            else:
                day = now_ist().date()

            # Sessions on this IST date; push branch/name filters into SQL.
            sess_q = db.query(LoginSession).filter(LoginSession.session_date == day)
            if branch_code or search:
                sess_q = sess_q.join(User, User.user_id == LoginSession.user_id)
                if branch_code:
                    sess_q = sess_q.filter(User.branch == branch_code)
                if search:
                    like = f"%{search.strip()}%"
                    sess_q = sess_q.filter(
                        or_(LoginSession.user_id.like(like), User.name.like(like))
                    )
            sessions = sess_q.order_by(LoginSession.login_time.desc()).all()

            # Users for branch + name display
            user_ids = list({s.user_id for s in sessions})
            user_map = {}
            if user_ids:
                for u in db.query(User).filter(User.user_id.in_(user_ids)).all():
                    user_map[u.user_id] = u

            now = now_ist()
            today = now.date()
            MAX_OPEN_SESSION_SECS = 10 * 3600  # cap a forgotten open tab at 10h

            # Login times per user (sorted) — an open session can't be Active
            # if the same user logged in again later; it was superseded.
            logins_by_user = {}
            for s in sessions:
                if s.login_time:
                    logins_by_user.setdefault(s.user_id, []).append(s.login_time)
            for lst in logins_by_user.values():
                lst.sort()

            def next_login_after(s):
                for t in logins_by_user.get(s.user_id, []):
                    if t > s.login_time:
                        return t
                return None

            def fmt_secs(secs):
                if not secs or secs < 0:
                    return "0h 0m"
                h = secs // 3600
                m = (secs % 3600) // 60
                return f"{h}h {m}m"

            # One place decides logout_time + status + duration.
            # The 3 client signals (manual/auto/close) are best-effort; on a
            # crash, power loss, dropped sendBeacon, or frozen background tab
            # NONE of them fire and logout_time stays NULL forever -> stuck on
            # "Active". So self-heal: any open session that can't still be live
            # is marked 'expired' with a synthesized logout time.
            def resolve_session(s):
                # Already closed -> trust stored values.
                if s.logout_time:
                    if s.duration_seconds is not None:
                        dur = max(s.duration_seconds, 0)
                    else:
                        dur = max(int((s.logout_time - s.login_time).total_seconds()), 0)
                    return s.logout_time, (s.logout_type or "unknown"), dur

                if not s.login_time:
                    return None, "no-logout", 0

                # Superseded: the user logged in again after this open session,
                # so it can't still be live — close it at that next login.
                nxt = next_login_after(s)
                if nxt:
                    dur = int((nxt - s.login_time).total_seconds())
                    if dur > MAX_OPEN_SESSION_SECS:
                        nxt = s.login_time + timedelta(seconds=MAX_OPEN_SESSION_SECS)
                        dur = MAX_OPEN_SESSION_SECS
                    return nxt, "relogin", max(dur, 0)

                # Open session on TODAY's date.
                if day == today:
                    open_secs = int((now - s.login_time).total_seconds())
                    if open_secs <= MAX_OPEN_SESSION_SECS:
                        return None, "active", max(open_secs, 0)      # genuinely live
                    synth = s.login_time + timedelta(seconds=MAX_OPEN_SESSION_SECS)
                    return synth, "expired", MAX_OPEN_SESSION_SECS

                # Open session on a PAST day -> never closed. Auto-close at
                # min(login + 10h, end of that day) so it never shows as Active.
                end_of_day = datetime.combine(day, datetime.max.time())
                close_at = min(s.login_time + timedelta(seconds=MAX_OPEN_SESSION_SECS), end_of_day)
                dur = max(int((close_at - s.login_time).total_seconds()), 0)
                return close_at, "expired", dur

            rows = []
            for s in sessions:
                u = user_map.get(s.user_id)
                branch = u.branch if u else (s.branch or 'N/A')
                user_name = (u.name if u else None) or s.user_name or 'N/A'

                logout_dt, status_label, secs = resolve_session(s)

                rows.append({
                    "user_id": s.user_id,
                    "user_name": user_name,
                    "branch": branch,
                    "branch_display": get_branch_display_name(branch),
                    "login_time": s.login_time.isoformat() if s.login_time else None,
                    "logout_time": logout_dt.isoformat() if logout_dt else None,
                    "session_seconds": secs,
                    "work_time": fmt_secs(secs),
                    "logout_type": status_label,
                })

            return {"date": day.strftime('%Y-%m-%d'), "total_sessions": len(rows), "rows": rows}

        except Exception as e:
            print(f"Error in get_employee_time_report: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"date": target_date or "", "total_sessions": 0, "rows": []}                        

    @staticmethod
    async def get_all_branches_letter_counts(db: Session):
        """
        Letter status counts grouped by the SENDER's branch, for ALL branches at once.
        ONE join (LetterSendRecord -> User on sent_by_id) — no per-branch round trips.
        Returns { branch_code: {total, sent, draft, failed} }.
        """
        try:
            from app.models.letter_model import LetterSendRecord
            from app.models.user_model import User

            rows = db.query(
                User.branch.label('branch'),
                LetterSendRecord.status.label('status'),
                func.count(LetterSendRecord.id).label('cnt')
            ).select_from(LetterSendRecord).join(
                User, User.user_id == LetterSendRecord.sent_by_id
            ).group_by(User.branch, LetterSendRecord.status).all()

            result = {}
            for branch, status_val, cnt in rows:
                if not branch:
                    continue
                c = int(cnt or 0)
                b = result.setdefault(branch, {"total": 0, "sent": 0, "draft": 0, "failed": 0})
                b["total"] += c
                s = (status_val or 'sent').lower()
                if s in b:
                    b[s] += c   # any other status (e.g. 'pending') still counts toward total
            return result
        except Exception as e:
            print(f"Error in get_all_branches_letter_counts: {str(e)}")
            import traceback
            traceback.print_exc()
            return {}

    @staticmethod
    async def get_branch_letter_records(db: Session, branch_code: str):
        """
        All letters sent by users whose PRIMARY branch is `branch_code`.
        PERF: light columns ONLY (no letter_html / letter_body / letter_fields blobs).
        Customer name/phone batched (chunked IN()); sender name resolved from users.
        """
        try:
            from app.models.letter_model import LetterSendRecord
            from app.models.user_model import User
            from app.models.customer_model import Customer

            # Users in this branch (id -> name) in ONE query
            branch_users = db.query(User.user_id, User.name).filter(
                User.branch == branch_code
            ).all()
            if not branch_users:
                return {"total": 0, "sent": 0, "draft": 0, "letters": []}

            user_name_map = {u.user_id: u.name for u in branch_users}
            branch_user_ids = list(user_name_map.keys())   # a branch has few users -> safe IN()

            # PERF: the `attachments` JSON column stores full base64 file CONTENT —
            # selecting it drags megabytes per letter across the wire just to show
            # file names. Extract the names inside SQL Server instead (OPENJSON needs
            # 2016+, STRING_AGG 2017+); if that fails, fall back to the old column.
            base_cols = [
                LetterSendRecord.id,
                LetterSendRecord.ref_no,
                LetterSendRecord.instance_id,
                LetterSendRecord.customer_id,
                LetterSendRecord.format_type_name,
                LetterSendRecord.subject,
                LetterSendRecord.channels,
                LetterSendRecord.sent_email,
                LetterSendRecord.sent_whatsapp,
                LetterSendRecord.email_to,
                LetterSendRecord.email_cc,
                LetterSendRecord.whatsapp_to,
                LetterSendRecord.status,
                LetterSendRecord.sent_by_id,
                LetterSendRecord.sent_by_name,
                LetterSendRecord.created_at,
            ]
            from sqlalchemy import literal_column
            names_expr = literal_column(
                "(SELECT STRING_AGG(COALESCE(JSON_VALUE(a.[value], '$.name'), "
                "CASE WHEN a.[type] = 1 THEN a.[value] END), '||') "
                "FROM OPENJSON(letter_send_records.attachments) a)"
            ).label('attachment_names')
            fast_names = True
            try:
                rows = db.query(*base_cols, names_expr).filter(
                    LetterSendRecord.sent_by_id.in_(branch_user_ids)
                ).order_by(LetterSendRecord.created_at.desc()).all()
            except Exception:
                db.rollback()
                fast_names = False
                rows = db.query(*base_cols, LetterSendRecord.attachments).filter(
                    LetterSendRecord.sent_by_id.in_(branch_user_ids)
                ).order_by(LetterSendRecord.created_at.desc()).all()

            if not rows:
                return {"total": 0, "sent": 0, "draft": 0, "letters": []}

            # Batch customer lookup — chunked to stay under SQL Server's 2100-param limit
            instance_ids = list({r.instance_id for r in rows if r.instance_id})
            customer_map = {}
            if instance_ids:
                CHUNK = 500
                for i in range(0, len(instance_ids), CHUNK):
                    part = instance_ids[i:i + CHUNK]
                    for c in db.query(
                        Customer.instance_id, Customer.customer_name,
                        Customer.phone_number, Customer.branch_id
                    ).filter(Customer.instance_id.in_(part)).all():
                        customer_map[c.instance_id] = c

            letters, sent_count, draft_count = [], 0, 0
            for idx, r in enumerate(rows, 1):
                cust = customer_map.get(r.instance_id)
                s = (r.status or 'sent').lower()
                if s == 'sent':
                    sent_count += 1
                elif s == 'draft':
                    draft_count += 1

                # File NAMES only — never the heavy base64 content
                if fast_names:
                    attachment_names = [n for n in (r.attachment_names or '').split('||') if n]
                else:
                    attachment_names = [
                        (a.get("name") if isinstance(a, dict) else a)
                        for a in (r.attachments or [])
                        if (a.get("name") if isinstance(a, dict) else a)
                    ]

                letters.append({
                    "s_no": idx,
                    "id": r.id,
                    "ref_no": r.ref_no or "-",
                    "instance_id": r.instance_id or "-",
                    "customer_id": r.customer_id,
                    "customer_name": cust.customer_name if cust else "-",
                    "phone_number": cust.phone_number if cust else "-",
                    "branch_id": cust.branch_id if cust else "-",
                    "format_type_name": r.format_type_name or "-",
                    "subject": r.subject or "-",
                    "channels": r.channels or [],
                    "sent_email": bool(r.sent_email),
                    "sent_whatsapp": bool(r.sent_whatsapp),
                    "email_to": r.email_to or "-",
                    "email_cc": r.email_cc or "-",
                    "whatsapp_to": r.whatsapp_to or "-",
                    "attachment_names": attachment_names,
                    "status": r.status or "-",
                    "sent_by_id": r.sent_by_id or "-",
                    "sent_by_name": user_name_map.get(r.sent_by_id) or r.sent_by_name or "-",
                    "created_at": r.created_at.isoformat() if r.created_at else None,
                })

            return {"total": len(letters), "sent": sent_count, "draft": draft_count, "letters": letters}

        except Exception as e:
            print(f"Error in get_branch_letter_records: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"total": 0, "sent": 0, "draft": 0, "letters": []}

    @staticmethod
    async def get_branch_letter_html(db: Session, branch_code: str, letter_id: int):
        """
        Return ONE letter's stored HTML, but ONLY if its sender belongs to `branch_code`
        (lets a branch/master admin view letters from their branch; blocks other branches).
        The frontend renders this HTML to a banded PDF.
        """
        try:
            from app.models.letter_model import LetterSendRecord
            from app.models.user_model import User

            rec = db.query(
                LetterSendRecord.ref_no,
                LetterSendRecord.subject,
                LetterSendRecord.letter_html,
                LetterSendRecord.letter_body,
                LetterSendRecord.sent_by_id,
            ).filter(LetterSendRecord.id == letter_id).first()

            if not rec:
                return None

            # Branch ownership check — sender must be in this branch
            in_branch = db.query(User.user_id).filter(
                User.user_id == rec.sent_by_id,
                User.branch == branch_code
            ).first()
            if not in_branch:
                return None

            html = rec.letter_html
            if not html or not str(html).strip():
                body = (rec.letter_body or '').replace('\n', '<br/>')
                html = f"<div style='font-family:Arial;padding:24px;'><h3>{rec.subject or ''}</h3><div>{body}</div></div>"

            return {
                "ref_no": rec.ref_no or f"letter_{letter_id}",
                "subject": rec.subject or "",
                "letter_html": str(html),
            }
        except Exception as e:
            print(f"Error in get_branch_letter_html: {str(e)}")
            import traceback
            traceback.print_exc()
            return None            